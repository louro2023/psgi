from __future__ import annotations

import csv
import io
import json
import mimetypes
import os
import secrets
import shutil
import sqlite3
import time
import uuid
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any

from flask import (
    Flask,
    Response,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
UPLOAD_DIR = BASE_DIR / "uploads"
BACKUP_DIR = BASE_DIR / "backups"
DATABASE = INSTANCE_DIR / "pgsi.sqlite3"

INSTANCE_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {
    "csv",
    "doc",
    "docx",
    "gif",
    "jpeg",
    "jpg",
    "pdf",
    "png",
    "txt",
    "xls",
    "xlsm",
    "xlsx",
}

APP_VERSION = "1.0.0"
TODAY = date.today()


app = Flask(__name__)

_secret_key = os.environ.get("PGSI_SECRET_KEY")
if not _secret_key:
    _secret_key = secrets.token_hex(32)
    print(
        "WARNING: PGSI_SECRET_KEY not set. Using random key — sessions will not survive restarts.\n"
        "         Set PGSI_SECRET_KEY env var before deploying to production."
    )

app.config.update(
    SECRET_KEY=_secret_key,
    MAX_CONTENT_LENGTH=25 * 1024 * 1024,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=os.environ.get("PGSI_SECURE_COOKIES", "0") == "1",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
)

from werkzeug.middleware.proxy_fix import ProxyFix  # noqa: E402
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS perfis (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL UNIQUE,
    descricao TEXT,
    permissoes TEXT NOT NULL DEFAULT '{}',
    status TEXT NOT NULL DEFAULT 'Ativo',
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS usuarios (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    email TEXT NOT NULL UNIQUE,
    senha_hash TEXT NOT NULL,
    perfil_id INTEGER NOT NULL,
    setor TEXT,
    status TEXT NOT NULL DEFAULT 'Ativo',
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (perfil_id) REFERENCES perfis(id)
);

CREATE TABLE IF NOT EXISTS pessoas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    cpf TEXT,
    matricula TEXT,
    email TEXT,
    telefone TEXT,
    cargo TEXT,
    setor TEXT,
    perfil_acesso TEXT,
    unidade_administrativa TEXT,
    status TEXT NOT NULL DEFAULT 'Ativo',
    observacoes TEXT,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS ativos_inventario (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    identificador TEXT,
    tipo TEXT,
    responsavel_id INTEGER,
    criticidade TEXT,
    status TEXT NOT NULL DEFAULT 'Ativo',
    localizacao TEXT,
    unidade TEXT,
    data_aquisicao TEXT,
    data_revisao TEXT,
    observacoes TEXT,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    ativo INTEGER NOT NULL DEFAULT 1,
    FOREIGN KEY (responsavel_id) REFERENCES pessoas(id)
);

CREATE TABLE IF NOT EXISTS controles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    titulo TEXT NOT NULL,
    descricao TEXT,
    categoria TEXT,
    tipo TEXT,
    eixo TEXT,
    area_responsavel TEXT,
    responsavel_id INTEGER NOT NULL,
    periodicidade TEXT,
    status TEXT NOT NULL,
    criticidade TEXT,
    data_criacao TEXT NOT NULL,
    data_atualizacao TEXT,
    data_revisao TEXT,
    prazo TEXT,
    observacoes TEXT,
    ativo INTEGER NOT NULL DEFAULT 1,
    FOREIGN KEY (responsavel_id) REFERENCES pessoas(id)
);

CREATE TABLE IF NOT EXISTS riscos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    titulo TEXT NOT NULL,
    descricao TEXT,
    categoria TEXT,
    eixo TEXT,
    ativo_relacionado TEXT,
    controle_id INTEGER,
    probabilidade INTEGER NOT NULL,
    impacto INTEGER NOT NULL,
    pontuacao INTEGER NOT NULL,
    nivel_risco TEXT NOT NULL,
    responsavel_id INTEGER,
    plano_tratamento TEXT,
    prazo TEXT,
    status TEXT NOT NULL,
    observacoes TEXT,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    ativo INTEGER NOT NULL DEFAULT 1,
    FOREIGN KEY (controle_id) REFERENCES controles(id),
    FOREIGN KEY (responsavel_id) REFERENCES pessoas(id)
);

CREATE TABLE IF NOT EXISTS incidentes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    numero TEXT NOT NULL UNIQUE,
    titulo TEXT NOT NULL,
    descricao TEXT,
    tipo TEXT,
    gravidade TEXT,
    ativo_afetado TEXT,
    area_afetada TEXT,
    status TEXT NOT NULL,
    responsavel_id INTEGER,
    data_abertura TEXT NOT NULL,
    data_encerramento TEXT,
    causa TEXT,
    acoes_tomadas TEXT,
    licoes_aprendidas TEXT,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    ativo INTEGER NOT NULL DEFAULT 1,
    FOREIGN KEY (responsavel_id) REFERENCES pessoas(id)
);

CREATE TABLE IF NOT EXISTS acoes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    titulo TEXT NOT NULL,
    descricao TEXT,
    origem TEXT,
    eixo TEXT,
    prioridade TEXT NOT NULL,
    status TEXT NOT NULL,
    responsavel_id INTEGER,
    controle_id INTEGER,
    risco_id INTEGER,
    incidente_id INTEGER,
    data_inicio TEXT,
    prazo TEXT NOT NULL,
    data_conclusao TEXT,
    evidencia_conclusao TEXT,
    observacoes TEXT,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    ativo INTEGER NOT NULL DEFAULT 1,
    FOREIGN KEY (responsavel_id) REFERENCES pessoas(id),
    FOREIGN KEY (controle_id) REFERENCES controles(id),
    FOREIGN KEY (risco_id) REFERENCES riscos(id),
    FOREIGN KEY (incidente_id) REFERENCES incidentes(id)
);

CREATE TABLE IF NOT EXISTS arquivos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome_original TEXT NOT NULL,
    nome_salvo TEXT NOT NULL,
    tipo TEXT,
    caminho TEXT NOT NULL,
    tamanho INTEGER NOT NULL,
    classificacao TEXT NOT NULL,
    descricao TEXT,
    eixo TEXT,
    usuario_upload_id INTEGER,
    controle_id INTEGER,
    risco_id INTEGER,
    incidente_id INTEGER,
    acao_id INTEGER,
    pessoa_id INTEGER,
    relatorio_id INTEGER,
    extraido_texto TEXT,
    extraido_em TEXT,
    ativo INTEGER NOT NULL DEFAULT 1,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (usuario_upload_id) REFERENCES usuarios(id),
    FOREIGN KEY (controle_id) REFERENCES controles(id),
    FOREIGN KEY (risco_id) REFERENCES riscos(id),
    FOREIGN KEY (incidente_id) REFERENCES incidentes(id),
    FOREIGN KEY (acao_id) REFERENCES acoes(id),
    FOREIGN KEY (pessoa_id) REFERENCES pessoas(id)
);

CREATE TABLE IF NOT EXISTS indicadores (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    descricao TEXT,
    formula TEXT,
    periodicidade TEXT,
    valor TEXT,
    referencia TEXT,
    criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS logs_auditoria (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    usuario_id INTEGER,
    acao TEXT NOT NULL,
    tabela_afetada TEXT,
    registro_id INTEGER,
    data_hora TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    ip TEXT,
    detalhes TEXT,
    FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
);
"""


MODULE_LABELS = {
    "dashboard": "Dashboard",
    "usuarios": "Usuários",
    "perfis": "Perfis",
    "pessoas": "Pessoas",
    "controles": "Controles",
    "riscos": "Riscos",
    "incidentes": "Incidentes",
    "acoes": "Plano de ação",
    "arquivos": "Documentos",
    "relatorios": "Relatórios",
    "logs": "Auditoria",
    "backup": "Backup",
}

ACTION_LABELS = {
    "view": "Visualizar",
    "create": "Cadastrar",
    "edit": "Editar",
    "delete": "Inativar",
    "download": "Baixar",
    "analyze": "Analisar",
    "export": "Exportar",
}

PERMISSION_MATRIX = {
    "dashboard": ["view"],
    "usuarios": ["view", "create", "edit", "delete"],
    "perfis": ["view", "create", "edit", "delete"],
    "pessoas": ["view", "create", "edit", "delete"],
    "controles": ["view", "create", "edit", "delete"],
    "riscos": ["view", "create", "edit", "delete"],
    "incidentes": ["view", "create", "edit", "delete"],
    "acoes": ["view", "create", "edit", "delete"],
    "arquivos": ["view", "create", "edit", "delete", "download", "analyze"],
    "relatorios": ["view", "export"],
    "logs": ["view"],
    "backup": ["view", "create"],
}


def permissions_for_role(role: str) -> dict[str, list[str]]:
    all_permissions = {module: actions[:] for module, actions in PERMISSION_MATRIX.items()}
    if role == "Administrador":
        return all_permissions
    if role == "Gestor":
        return {
            "dashboard": ["view"],
            "pessoas": ["view", "create", "edit"],
            "controles": ["view", "create", "edit", "delete"],
            "riscos": ["view", "create", "edit", "delete"],
            "incidentes": ["view", "create", "edit"],
            "acoes": ["view", "create", "edit", "delete"],
            "arquivos": ["view", "create", "edit", "download", "analyze"],
            "relatorios": ["view", "export"],
            "logs": ["view"],
        }
    if role == "Técnico":
        return {
            "dashboard": ["view"],
            "pessoas": ["view"],
            "controles": ["view", "edit"],
            "riscos": ["view", "create", "edit"],
            "incidentes": ["view", "create", "edit"],
            "acoes": ["view", "edit"],
            "arquivos": ["view", "create", "download", "analyze"],
        }
    if role == "Auditor":
        return {
            "dashboard": ["view"],
            "pessoas": ["view"],
            "controles": ["view"],
            "riscos": ["view"],
            "incidentes": ["view"],
            "acoes": ["view"],
            "arquivos": ["view", "download", "analyze"],
            "relatorios": ["view", "export"],
            "logs": ["view"],
        }
    if role == "Consulta":
        return {
            "dashboard": ["view"],
            "pessoas": ["view"],
            "controles": ["view"],
            "riscos": ["view"],
            "incidentes": ["view"],
            "acoes": ["view"],
            "arquivos": ["view", "download"],
            "relatorios": ["view"],
        }
    return {
        "dashboard": ["view"],
        "arquivos": ["view", "create", "download"],
    }


CONTROL_STATUSES = [
    "Não iniciado",
    "Em andamento",
    "Implementado",
    "Parcialmente implementado",
    "Pendente",
    "Vencido",
    "Em revisão",
    "Concluído",
]

ACTION_STATUSES = [
    "Aberta",
    "Em andamento",
    "Aguardando validação",
    "Concluída",
    "Vencida",
    "Cancelada",
]

INCIDENT_STATUSES = ["Aberto", "Em andamento", "Contido", "Encerrado", "Cancelado"]

ASSET_TYPES = [
    "Servidor",
    "Notebook",
    "Desktop",
    "Dispositivo de rede",
    "Firewall",
    "Sistema",
    "Aplicação",
    "Banco de dados",
    "Serviço",
    "Informação/Dado",
    "Outro",
]

ASSET_CRITICALITIES = ["Baixa", "Média", "Alta", "Crítica"]

ASSET_STATUSES = ["Ativo", "Em manutenção", "Reservado", "Desativado", "Descartado"]

PGSI_CONTROLS = [
    ("Inventário e controle de ativos corporativos", "Gestão de Ativos", "Segurança", "Manter inventário atualizado dos ativos corporativos e responsáveis.", "DTI", "Mensal", "Alta"),
    ("Inventário e controle de ativos de software", "Gestão de Ativos", "Segurança", "Identificar softwares autorizados, suportados e exceções aprovadas.", "DTI", "Mensal", "Alta"),
    ("Proteção de dados", "Privacidade", "Privacidade", "Classificar, manusear, reter e descartar dados com segurança.", "Encarregado de Dados", "Anual", "Crítica"),
    ("Configuração segura de ativos corporativos e software", "Infraestrutura", "Segurança", "Aplicar configurações seguras em servidores, estações e aplicações.", "Infraestrutura", "Trimestral", "Alta"),
    ("Gestão de contas", "Controle de Acesso", "Segurança", "Gerenciar ciclo de vida de contas e privilégios.", "Suporte", "Mensal", "Alta"),
    ("Gestão do controle de acesso", "Controle de Acesso", "Conformidade", "Prevenir acesso não autorizado a sistemas, dados e serviços.", "Segurança da Informação", "Semestral", "Crítica"),
    ("Gestão contínua de vulnerabilidades", "Vulnerabilidades", "Risco", "Identificar, priorizar e tratar vulnerabilidades técnicas.", "Segurança da Informação", "Mensal", "Crítica"),
    ("Gestão de registros de auditoria", "Auditoria", "Conformidade", "Coletar, reter e revisar logs de eventos relevantes.", "Infraestrutura", "Semanal", "Alta"),
    ("Proteções de e-mail e navegador Web", "Proteção", "Segurança", "Reduzir exposição a phishing, domínios maliciosos e anexos perigosos.", "Redes", "Trimestral", "Alta"),
    ("Defesas contra malware", "Proteção", "Segurança", "Impedir ou controlar execução de códigos maliciosos.", "Suporte", "Mensal", "Crítica"),
    ("Recuperação de dados", "Continuidade", "Segurança", "Executar backups, proteger cópias e testar restaurações.", "Infraestrutura", "Trimestral", "Crítica"),
    ("Gestão de infraestrutura de redes", "Infraestrutura", "Segurança", "Rastrear, reportar e corrigir dispositivos de rede.", "Redes", "Semestral", "Alta"),
    ("Monitoramento e defesa da rede", "Monitoramento", "Segurança", "Operar processos e ferramentas de defesa de rede.", "Redes", "Contínua", "Crítica"),
    ("Conscientização sobre segurança e treinamento de competências", "Conscientização", "Melhoria Contínua", "Treinar usuários para práticas seguras e comunicação de incidentes.", "Gestão de Pessoas", "Anual", "Média"),
    ("Gestão de provedor de serviços", "Fornecedores", "Conformidade", "Avaliar e monitorar provedores com requisitos de segurança.", "Contratos", "Anual", "Alta"),
    ("Segurança de aplicações", "Aplicações", "Segurança", "Gerenciar ciclo de vida seguro de software adquirido ou desenvolvido.", "Sistemas", "Semestral", "Alta"),
    ("Gestão de respostas a incidentes", "Incidentes", "Segurança", "Definir funções, contatos, comunicação e análise pós-incidente.", "Segurança da Informação", "Anual", "Crítica"),
    ("Testes de invasão", "Vulnerabilidades", "Risco", "Testar resiliência por exploração controlada de fraquezas.", "Segurança da Informação", "Anual", "Alta"),
    ("Política de Segurança da Informação", "Governança", "Conformidade", "Manter política aprovada, divulgada e revisada.", "Comitê de Segurança", "Anual", "Alta"),
    ("Organização da segurança da informação", "Governança", "Conformidade", "Definir estrutura, papéis e responsabilidades de segurança.", "Comitê de Segurança", "Anual", "Alta"),
    ("Segurança em recursos humanos", "Pessoas", "Segurança", "Aplicar controles antes, durante e após vínculos de trabalho.", "Gestão de Pessoas", "Anual", "Média"),
    ("Segurança física e do ambiente", "Segurança Física", "Infraestrutura", "Proteger instalações, salas, perímetros e equipamentos críticos.", "Administração", "Anual", "Alta"),
    ("Segurança de equipamentos", "Ativos", "Infraestrutura", "Proteger equipamentos dentro e fora das dependências institucionais.", "Patrimônio", "Anual", "Média"),
    ("Gerenciamento das operações e comunicações", "Operações", "Segurança", "Controlar mudanças, rotinas operacionais e comunicações.", "DTI", "Semestral", "Alta"),
    ("Manuseio de mídias", "Proteção de Dados", "Privacidade", "Controlar uso, descarte e transporte de mídias.", "Suporte", "Anual", "Média"),
    ("Troca de informações", "Proteção de Dados", "Privacidade", "Formalizar e proteger trocas internas e externas de informações.", "Segurança da Informação", "Anual", "Alta"),
    ("Controle de acesso à rede", "Controle de Acesso", "Segurança", "Prevenir acesso não autorizado aos serviços de rede.", "Redes", "Semestral", "Crítica"),
]


SECURITY_AREAS = [
    {
        "key": "vulnerabilidades",
        "number": 1,
        "title": "Gestão de Vulnerabilidades",
        "summary": "Identifica, prioriza e acompanha fragilidades técnicas antes que elas sejam exploradas.",
        "explanation": "Centraliza varreduras, achados, severidade, prazos de correção, responsáveis e evidências de tratamento.",
        "objectives": ["Manter vulnerabilidades críticas sob tratamento", "Definir prioridade por risco", "Registrar evidências de correção"],
        "terms": ["vulnerabilidade", "vulnerabilidades", "teste de invasão", "invasão", "varredura", "patch", "atualização", "exploração"],
        "control_type": "Risco",
        "responsible_area": "Segurança da Informação",
    },
    {
        "key": "seguranca-fisica",
        "number": 2,
        "title": "Segurança Física e do Ambiente",
        "summary": "Protege instalações, salas, perímetros, energia, climatização e equipamentos críticos.",
        "explanation": "Organiza controles físicos e ambientais que reduzem danos, indisponibilidade e acesso indevido a ambientes sensíveis.",
        "objectives": ["Controlar acesso físico", "Proteger ambientes críticos", "Registrar revisões e evidências ambientais"],
        "terms": ["segurança física", "ambiente", "perímetro", "sala", "instalação", "equipamento", "energia", "climatização"],
        "control_type": "Infraestrutura",
        "responsible_area": "Administração",
    },
    {
        "key": "incidentes",
        "number": 3,
        "title": "Gestão de Incidentes de Segurança da Informação",
        "summary": "Registra, acompanha e documenta eventos que afetam confidencialidade, integridade ou disponibilidade.",
        "explanation": "Conecta incidentes com ações, causas prováveis, evidências, lições aprendidas e status de resposta.",
        "objectives": ["Reduzir tempo de resposta", "Documentar contenção e causa", "Gerar histórico de lições aprendidas"],
        "terms": ["incidente", "incidentes", "phishing", "malware", "vazamento", "indisponibilidade", "falha", "resposta"],
        "control_type": "Segurança",
        "responsible_area": "Segurança da Informação",
    },
    {
        "key": "ativos",
        "number": 4,
        "title": "Gestão de Ativos",
        "summary": "Mantém inventário, criticidade, responsáveis e ciclo de vida dos ativos corporativos.",
        "explanation": "Relaciona ativos de hardware, software, serviços, dados e responsáveis aos controles aplicáveis.",
        "objectives": ["Atualizar inventário", "Definir responsáveis por ativos", "Relacionar controles aos ativos críticos"],
        "terms": ["ativo", "ativos", "inventário", "patrimônio", "equipamento", "hardware", "corporativo"],
        "control_type": "Segurança",
        "responsible_area": "DTI",
    },
    {
        "key": "infraestrutura-rede",
        "number": 5,
        "title": "Gestão de Infraestrutura de Rede",
        "summary": "Organiza configuração, monitoramento e manutenção de redes, firewalls e conectividade.",
        "explanation": "Consolida controles sobre dispositivos de rede, segmentação, roteamento, regras e documentação técnica.",
        "objectives": ["Manter dispositivos rastreados", "Controlar configurações de rede", "Registrar mudanças e evidências"],
        "terms": ["rede", "redes", "firewall", "roteamento", "switch", "infraestrutura", "conectividade", "segmentação"],
        "control_type": "Infraestrutura",
        "responsible_area": "Redes",
    },
    {
        "key": "controle-acesso",
        "number": 6,
        "title": "Gestão de Controle de Acesso",
        "summary": "Previne acesso não autorizado a sistemas, dados, redes e serviços institucionais.",
        "explanation": "Acompanha contas, privilégios, revisões de acesso, autenticação, autorização e evidências de revisão.",
        "objectives": ["Revisar permissões periodicamente", "Tratar contas privilegiadas", "Reduzir acessos indevidos"],
        "terms": ["acesso", "conta", "contas", "permissão", "permissões", "privilégio", "autenticação", "autorização"],
        "control_type": "Conformidade",
        "responsible_area": "Segurança da Informação",
    },
    {
        "key": "software",
        "number": 7,
        "title": "Gestão de Software",
        "summary": "Controla softwares autorizados, versões suportadas, scripts, bibliotecas e aplicações.",
        "explanation": "Apoia o inventário de software, remoção de itens não autorizados e gestão segura de aplicações.",
        "objectives": ["Manter software autorizado", "Remover aplicações não aprovadas", "Acompanhar versões suportadas"],
        "terms": ["software", "aplicação", "aplicações", "biblioteca", "script", "versão", "autorizado"],
        "control_type": "Segurança",
        "responsible_area": "Sistemas",
    },
    {
        "key": "mudancas",
        "number": 8,
        "title": "Gestão de Mudanças",
        "summary": "Controla alterações em sistemas, infraestrutura, processos e serviços críticos.",
        "explanation": "Registra impactos, aprovações, janelas, responsáveis, evidências e riscos associados às mudanças.",
        "objectives": ["Reduzir mudanças sem aprovação", "Registrar impacto e rollback", "Conectar mudanças a riscos e ações"],
        "terms": ["mudança", "mudanças", "alteração", "alterações", "implantação", "rollback", "configuração"],
        "control_type": "Conformidade",
        "responsible_area": "DTI",
    },
    {
        "key": "fornecedores-terceiros",
        "number": 9,
        "title": "Gestão de Fornecedores e Terceiros",
        "summary": "Avalia provedores, contratos, acessos externos e requisitos de segurança.",
        "explanation": "Mantém evidências de avaliação, classificação, monitoramento e encerramento seguro de fornecedores.",
        "objectives": ["Classificar fornecedores", "Incluir requisitos de segurança em contratos", "Monitorar prestadores críticos"],
        "terms": ["fornecedor", "fornecedores", "terceiro", "terceiros", "provedor", "provedores", "contrato", "contratos"],
        "control_type": "Conformidade",
        "responsible_area": "Contratos",
    },
    {
        "key": "continuidade-negocios",
        "number": 10,
        "title": "Gestão de Continuidade de Negócios",
        "summary": "Planeja a manutenção ou retomada de processos críticos diante de interrupções.",
        "explanation": "Conecta continuidade, impacto, prioridades, planos, testes e ações preventivas.",
        "objectives": ["Definir processos críticos", "Testar planos de continuidade", "Acompanhar ações de resiliência"],
        "terms": ["continuidade", "negócios", "resiliência", "interrupção", "processo crítico", "indisponibilidade"],
        "control_type": "Segurança",
        "responsible_area": "Governança",
    },
    {
        "key": "recuperacao-dados",
        "number": 11,
        "title": "Gestão de Recuperação de Dados",
        "summary": "Acompanha backups, restauração, retenção, proteção e testes de recuperação.",
        "explanation": "Organiza evidências de backup, testes de restore, isolamento das cópias e prazos de recuperação.",
        "objectives": ["Executar backups automatizados", "Testar restaurações", "Proteger dados de recuperação"],
        "terms": ["backup", "backups", "recuperação", "restauração", "restore", "retenção", "cópia"],
        "control_type": "Segurança",
        "responsible_area": "Infraestrutura",
    },
    {
        "key": "riscos-si",
        "number": 12,
        "title": "Gestão de Riscos de Segurança da Informação",
        "summary": "Identifica riscos, calcula nível por probabilidade e impacto e acompanha tratamento.",
        "explanation": "Mantém riscos vinculados a controles, responsáveis, prazos, planos de tratamento e evidências.",
        "objectives": ["Classificar riscos de forma consistente", "Tratar riscos altos", "Monitorar prazos e responsáveis"],
        "terms": ["risco", "riscos", "probabilidade", "impacto", "tratamento", "ameaça", "ameaças"],
        "control_type": "Risco",
        "responsible_area": "Segurança da Informação",
    },
    {
        "key": "logs-auditoria",
        "number": 13,
        "title": "Gestão de Registros (Logs) de Auditoria",
        "summary": "Define coleta, retenção, revisão e análise de registros de eventos relevantes.",
        "explanation": "Apoia rastreabilidade, investigação, conformidade e detecção de eventos suspeitos.",
        "objectives": ["Centralizar logs relevantes", "Revisar eventos críticos", "Manter trilha de auditoria"],
        "terms": ["log", "logs", "registro", "registros", "auditoria", "evento", "eventos", "rastreabilidade"],
        "control_type": "Conformidade",
        "responsible_area": "Infraestrutura",
    },
    {
        "key": "auditoria-conformidade",
        "number": 14,
        "title": "Gestão da Auditoria e Conformidade",
        "summary": "Acompanha requisitos, evidências, auditorias, revisões e aderência regulatória.",
        "explanation": "Organiza verificações de conformidade e relatórios para tomada de decisão e melhoria contínua.",
        "objectives": ["Mapear requisitos aplicáveis", "Registrar evidências de conformidade", "Acompanhar auditorias"],
        "terms": ["auditoria", "conformidade", "regulatório", "regulatória", "requisito", "revisão", "norma", "ISO"],
        "control_type": "Conformidade",
        "responsible_area": "Controle Interno",
    },
    {
        "key": "conscientizacao-treinamento",
        "number": 15,
        "title": "Conscientização, Educação e Treinamento",
        "summary": "Planeja campanhas, treinamentos e comunicação para reduzir riscos humanos.",
        "explanation": "Acompanha ações educativas, públicos, frequência, evidências e resultados de conscientização.",
        "objectives": ["Executar campanhas periódicas", "Treinar perfis críticos", "Registrar presença e materiais"],
        "terms": ["conscientização", "educação", "treinamento", "campanha", "usuário", "usuários", "competência"],
        "control_type": "Melhoria Contínua",
        "responsible_area": "Gestão de Pessoas",
    },
    {
        "key": "dados",
        "number": 16,
        "title": "Gestão de Dados",
        "summary": "Classifica, protege, retém, compartilha e descarta dados em todo o ciclo de vida.",
        "explanation": "Integra privacidade, proteção de dados pessoais, dados sensíveis, classificação e troca segura de informações.",
        "objectives": ["Classificar informações", "Proteger dados sensíveis", "Controlar retenção, descarte e compartilhamento"],
        "terms": ["dados", "dado", "privacidade", "sensível", "sensíveis", "classificação", "LGPD", "informação", "mídias"],
        "control_type": "Privacidade",
        "responsible_area": "Encarregado de Dados",
    },
]


AREA_PLAYBOOKS = {
    "vulnerabilidades": {
        "records": ["Vulnerabilidade identificada", "Sistema ou ativo afetado", "Severidade e prazo de correção", "Evidência de correção ou exceção aceita"],
        "evidence": ["Relatório de varredura", "Registro de patch aplicado", "Print da ferramenta de vulnerabilidade", "Termo de aceite de risco"],
        "workflow": ["Registrar a vulnerabilidade e o ativo afetado", "Classificar risco por probabilidade e impacto", "Criar ação de tratamento para itens altos", "Anexar evidência da correção"],
        "questions": ["Qual ativo está vulnerável?", "Existe exploração conhecida?", "Quem validou a correção?"],
    },
    "seguranca-fisica": {
        "records": ["Ambiente ou sala protegida", "Controle de entrada física", "Ameaça ambiental", "Revisão de perímetro"],
        "evidence": ["Lista de acesso físico", "Fotos ou laudos de sala segura", "Registro de manutenção", "Relatório de alarme, energia ou climatização"],
        "workflow": ["Cadastrar controle físico", "Definir responsável pela instalação", "Registrar risco ambiental", "Anexar evidências de inspeção"],
        "questions": ["O ambiente possui controle de acesso?", "Há proteção contra energia, incêndio ou climatização?", "Quando foi a última inspeção?"],
    },
    "incidentes": {
        "records": ["Incidente reportado", "Causa provável", "Contenção e resposta", "Lições aprendidas"],
        "evidence": ["Relatório de incidente", "Linha do tempo da resposta", "Comunicação enviada", "Evidências técnicas da contenção"],
        "workflow": ["Registrar incidente no módulo próprio", "Vincular riscos e ações de resposta", "Documentar causa e contenção", "Fechar com lições aprendidas"],
        "questions": ["Qual serviço ou dado foi afetado?", "O incidente ainda está aberto?", "Que ação evita recorrência?"],
    },
    "ativos": {
        "records": ["Ativo corporativo", "Responsável pelo ativo", "Criticidade", "Ciclo de vida"],
        "evidence": ["Inventário de ativos", "Termo de responsabilidade", "Registro patrimonial", "Relatório de descoberta de ativos"],
        "workflow": ["Cadastrar controle de inventário", "Associar responsável", "Classificar criticidade", "Registrar riscos e evidências"],
        "questions": ["O ativo tem dono definido?", "Ele é crítico para o negócio?", "Está coberto por controles?"],
    },
    "infraestrutura-rede": {
        "records": ["Dispositivo de rede", "Configuração segura", "Regra de firewall", "Diagrama ou segmentação"],
        "evidence": ["Backup de configuração", "Diagrama de rede", "Relatório de regras", "Registro de mudança"],
        "workflow": ["Mapear infraestrutura", "Registrar controle de configuração", "Vincular mudanças e riscos", "Anexar evidências técnicas"],
        "questions": ["A configuração foi revisada?", "Há segmentação adequada?", "A mudança tem aprovação?"],
    },
    "controle-acesso": {
        "records": ["Conta ou perfil de acesso", "Privilégio concedido", "Revisão periódica", "Acesso excepcional"],
        "evidence": ["Relatório de usuários", "Ata de revisão de acessos", "Solicitação aprovada", "Print de perfil ou grupo"],
        "workflow": ["Cadastrar controle de acesso", "Definir revisão e responsável", "Registrar risco de privilégio", "Anexar evidência de revisão"],
        "questions": ["Quem aprovou o acesso?", "O privilégio ainda é necessário?", "Existe MFA ou controle compensatório?"],
    },
    "software": {
        "records": ["Software autorizado", "Versão suportada", "Aplicação não autorizada", "Biblioteca ou script crítico"],
        "evidence": ["Inventário de software", "Lista de softwares permitidos", "Registro de remoção", "Comprovante de atualização"],
        "workflow": ["Registrar inventário de software", "Classificar autorização", "Criar ação para remover itens não aprovados", "Anexar evidência"],
        "questions": ["O software é autorizado?", "A versão recebe suporte?", "Há exceção documentada?"],
    },
    "mudancas": {
        "records": ["Mudança planejada", "Impacto esperado", "Aprovação", "Rollback"],
        "evidence": ["Solicitação de mudança", "Plano de rollback", "Registro de aprovação", "Relatório pós-implantação"],
        "workflow": ["Registrar controle de mudança", "Avaliar risco e impacto", "Criar ações de execução", "Anexar validação pós-mudança"],
        "questions": ["A mudança foi aprovada?", "Há janela e rollback?", "Qual controle foi afetado?"],
    },
    "fornecedores-terceiros": {
        "records": ["Fornecedor crítico", "Contrato com requisito de segurança", "Avaliação periódica", "Acesso de terceiro"],
        "evidence": ["Contrato", "Checklist de avaliação", "SLA ou termo de confidencialidade", "Registro de encerramento de acesso"],
        "workflow": ["Cadastrar controle de fornecedores", "Classificar criticidade", "Registrar riscos contratuais", "Anexar avaliação"],
        "questions": ["O fornecedor trata dados sensíveis?", "Há cláusulas de segurança?", "O acesso foi revisado?"],
    },
    "continuidade-negocios": {
        "records": ["Processo crítico", "Plano de continuidade", "Tempo de recuperação", "Teste de continuidade"],
        "evidence": ["Plano de continuidade", "Relatório de teste", "Análise de impacto", "Ata de validação"],
        "workflow": ["Mapear processo crítico", "Definir controle de continuidade", "Criar ações de teste", "Anexar resultado"],
        "questions": ["Qual processo precisa continuar?", "O plano foi testado?", "O tempo de recuperação é aceitável?"],
    },
    "recuperacao-dados": {
        "records": ["Rotina de backup", "Teste de restauração", "Retenção", "Cópia isolada"],
        "evidence": ["Log de backup", "Relatório de restore", "Política de retenção", "Comprovante de cópia protegida"],
        "workflow": ["Cadastrar controle de backup", "Definir periodicidade", "Registrar teste de recuperação", "Anexar evidência"],
        "questions": ["Backup executou com sucesso?", "Restore foi testado?", "Cópias estão protegidas?"],
    },
    "riscos-si": {
        "records": ["Risco identificado", "Ameaça e vulnerabilidade", "Plano de tratamento", "Aceite ou mitigação"],
        "evidence": ["Matriz de risco", "Plano de tratamento", "Registro de aceite", "Relatório de monitoramento"],
        "workflow": ["Cadastrar risco", "Calcular probabilidade x impacto", "Criar ação para risco alto", "Monitorar prazo"],
        "questions": ["Qual ameaça gera o risco?", "O risco alto tem plano?", "Quem acompanha o tratamento?"],
    },
    "logs-auditoria": {
        "records": ["Fonte de log", "Evento crítico", "Retenção", "Revisão de auditoria"],
        "evidence": ["Relatório de logs", "Configuração de retenção", "Registro de revisão", "Alerta investigado"],
        "workflow": ["Cadastrar controle de logs", "Definir eventos relevantes", "Registrar revisão", "Anexar evidência"],
        "questions": ["Quais eventos são coletados?", "Há retenção suficiente?", "Quem revisa os alertas?"],
    },
    "auditoria-conformidade": {
        "records": ["Requisito aplicável", "Achado de auditoria", "Plano de adequação", "Evidência de conformidade"],
        "evidence": ["Relatório de auditoria", "Checklist ISO/CIS/LGPD", "Plano de correção", "Documento normativo"],
        "workflow": ["Registrar requisito ou achado", "Vincular controle e risco", "Criar plano de ação", "Gerar relatório"],
        "questions": ["Qual requisito está sendo atendido?", "Há evidência suficiente?", "Existe pendência de auditoria?"],
    },
    "conscientizacao-treinamento": {
        "records": ["Campanha de conscientização", "Treinamento realizado", "Público-alvo", "Resultado ou presença"],
        "evidence": ["Lista de presença", "Material de campanha", "Certificado", "Relatório de participação"],
        "workflow": ["Planejar treinamento", "Registrar controle de conscientização", "Criar ação para público pendente", "Anexar presença"],
        "questions": ["Quem precisa ser treinado?", "A campanha foi registrada?", "Há evidência de participação?"],
    },
    "dados": {
        "records": ["Categoria de dado", "Classificação da informação", "Base legal ou finalidade", "Retenção e descarte"],
        "evidence": ["Inventário de dados", "Tabela de temporalidade", "Registro de tratamento", "Política de classificação"],
        "workflow": ["Classificar dados", "Definir controles de proteção", "Registrar riscos de privacidade", "Anexar documentação"],
        "questions": ["Há dados pessoais ou sensíveis?", "A informação está classificada?", "Existe regra de retenção?"],
    },
}


ENTITY_CONFIGS: dict[str, dict[str, Any]] = {
    "pessoas": {
        "title": "Pessoas",
        "table": "pessoas",
        "module": "pessoas",
        "list_query": "SELECT p.* FROM pessoas p WHERE 1=1",
        "id_column": "p.id",
        "search_fields": ["p.nome", "p.email", "p.setor", "p.cargo", "p.matricula"],
        "filters": {"status": "p.status", "setor": "p.setor"},
        "columns": [
            ("nome", "Nome"),
            ("email", "E-mail"),
            ("cargo", "Cargo"),
            ("setor", "Setor"),
            ("perfil_acesso", "Perfil"),
            ("status", "Status"),
        ],
        "fields": [
            {"name": "nome", "label": "Nome completo", "type": "text", "required": True},
            {"name": "cpf", "label": "CPF", "type": "text"},
            {"name": "matricula", "label": "Matrícula", "type": "text"},
            {"name": "email", "label": "E-mail", "type": "email"},
            {"name": "telefone", "label": "Telefone", "type": "text"},
            {"name": "cargo", "label": "Cargo", "type": "text"},
            {"name": "setor", "label": "Setor", "type": "text"},
            {"name": "perfil_acesso", "label": "Perfil de acesso", "type": "select", "options": ["Administrador", "Gestor", "Técnico", "Auditor", "Consulta", "Usuário comum"]},
            {"name": "unidade_administrativa", "label": "Unidade administrativa", "type": "text"},
            {"name": "status", "label": "Status", "type": "select", "options": ["Ativo", "Inativo"], "required": True},
            {"name": "observacoes", "label": "Observações", "type": "textarea", "span": 2},
        ],
    },
    "controles": {
        "title": "Controles",
        "table": "controles",
        "module": "controles",
        "list_query": """
            SELECT c.*, p.nome AS responsavel_nome,
                   (SELECT COUNT(*) FROM arquivos a WHERE a.controle_id = c.id AND a.ativo = 1) AS evidencias
            FROM controles c
            LEFT JOIN pessoas p ON p.id = c.responsavel_id
            WHERE c.ativo = 1
        """,
        "id_column": "c.id",
        "search_fields": ["c.titulo", "c.descricao", "c.categoria", "c.area_responsavel", "p.nome"],
        "filters": {"status": "c.status", "categoria": "c.categoria", "responsavel_id": "c.responsavel_id", "setor": "c.area_responsavel"},
        "columns": [
            ("titulo", "Controle"),
            ("categoria", "Categoria"),
            ("tipo", "Tipo"),
            ("responsavel_nome", "Responsável"),
            ("periodicidade", "Periodicidade"),
            ("status", "Status"),
            ("prazo", "Prazo"),
            ("evidencias", "Evidências"),
        ],
        "fields": [
            {"name": "titulo", "label": "Nome do controle", "type": "text", "required": True, "span": 2},
            {"name": "categoria", "label": "Categoria", "type": "text", "required": True},
            {"name": "tipo", "label": "Tipo", "type": "select", "options": ["Segurança", "Privacidade", "Risco", "Conformidade", "Infraestrutura", "Melhoria Contínua"], "required": True},
            {"name": "descricao", "label": "Descrição", "type": "textarea", "span": 2},
            {"name": "area_responsavel", "label": "Área responsável", "type": "text"},
            {"name": "responsavel_id", "label": "Pessoa responsável", "type": "choice", "choice": "pessoas", "required": True},
            {"name": "periodicidade", "label": "Periodicidade", "type": "select", "options": ["Contínua", "Diária", "Semanal", "Mensal", "Trimestral", "Semestral", "Anual"], "required": True},
            {"name": "status", "label": "Status", "type": "select", "options": CONTROL_STATUSES, "required": True},
            {"name": "criticidade", "label": "Criticidade", "type": "select", "options": ["Baixa", "Média", "Alta", "Crítica"]},
            {"name": "data_criacao", "label": "Data de criação", "type": "date", "required": True},
            {"name": "data_atualizacao", "label": "Última atualização", "type": "date"},
            {"name": "data_revisao", "label": "Data prevista para revisão", "type": "date"},
            {"name": "prazo", "label": "Prazo", "type": "date"},
            {"name": "observacoes", "label": "Observações", "type": "textarea", "span": 2},
        ],
    },
    "riscos": {
        "title": "Riscos",
        "table": "riscos",
        "module": "riscos",
        "list_query": """
            SELECT r.*, p.nome AS responsavel_nome, c.titulo AS controle_titulo
            FROM riscos r
            LEFT JOIN pessoas p ON p.id = r.responsavel_id
            LEFT JOIN controles c ON c.id = r.controle_id
            WHERE r.ativo = 1
        """,
        "id_column": "r.id",
        "search_fields": ["r.titulo", "r.descricao", "r.categoria", "r.ativo_relacionado", "p.nome", "c.titulo"],
        "filters": {"status": "r.status", "categoria": "r.categoria", "responsavel_id": "r.responsavel_id"},
        "columns": [
            ("titulo", "Risco"),
            ("categoria", "Categoria"),
            ("controle_titulo", "Controle"),
            ("probabilidade", "Prob."),
            ("impacto", "Impacto"),
            ("nivel_risco", "Nível"),
            ("responsavel_nome", "Responsável"),
            ("prazo", "Prazo"),
            ("status", "Status"),
        ],
        "fields": [
            {"name": "titulo", "label": "Título do risco", "type": "text", "required": True, "span": 2},
            {"name": "descricao", "label": "Descrição", "type": "textarea", "span": 2},
            {"name": "categoria", "label": "Categoria", "type": "text"},
            {"name": "ativo_relacionado", "label": "Ativo relacionado", "type": "text"},
            {"name": "controle_id", "label": "Controle relacionado", "type": "choice", "choice": "controles"},
            {"name": "responsavel_id", "label": "Responsável", "type": "choice", "choice": "pessoas"},
            {"name": "probabilidade", "label": "Probabilidade", "type": "select", "options": [("1", "Baixa"), ("2", "Média"), ("3", "Alta")], "required": True},
            {"name": "impacto", "label": "Impacto", "type": "select", "options": [("1", "Baixo"), ("2", "Médio"), ("3", "Alto")], "required": True},
            {"name": "plano_tratamento", "label": "Plano de tratamento", "type": "textarea", "span": 2},
            {"name": "prazo", "label": "Prazo", "type": "date"},
            {"name": "status", "label": "Status", "type": "select", "options": ["Identificado", "Em tratamento", "Aceito", "Mitigado", "Monitorado", "Encerrado"], "required": True},
            {"name": "observacoes", "label": "Observações", "type": "textarea", "span": 2},
        ],
    },
    "incidentes": {
        "title": "Incidentes",
        "table": "incidentes",
        "module": "incidentes",
        "list_query": """
            SELECT i.*, p.nome AS responsavel_nome
            FROM incidentes i
            LEFT JOIN pessoas p ON p.id = i.responsavel_id
            WHERE i.ativo = 1
        """,
        "id_column": "i.id",
        "search_fields": ["i.numero", "i.titulo", "i.tipo", "i.ativo_afetado", "i.area_afetada", "p.nome"],
        "filters": {"status": "i.status", "responsavel_id": "i.responsavel_id", "setor": "i.area_afetada"},
        "columns": [
            ("numero", "Número"),
            ("titulo", "Incidente"),
            ("tipo", "Tipo"),
            ("gravidade", "Gravidade"),
            ("responsavel_nome", "Responsável"),
            ("data_abertura", "Abertura"),
            ("status", "Status"),
        ],
        "fields": [
            {"name": "numero", "label": "Número do incidente", "type": "text", "required": True},
            {"name": "titulo", "label": "Título", "type": "text", "required": True},
            {"name": "descricao", "label": "Descrição", "type": "textarea", "span": 2},
            {"name": "tipo", "label": "Tipo de incidente", "type": "select", "options": ["Indisponibilidade de sistema", "Vazamento de dados", "Acesso indevido", "Malware", "Phishing", "Falha de backup", "Falha de rede", "Perda de equipamento", "Outro"]},
            {"name": "gravidade", "label": "Gravidade", "type": "select", "options": ["Baixa", "Média", "Alta", "Crítica"]},
            {"name": "ativo_afetado", "label": "Sistema ou ativo afetado", "type": "text"},
            {"name": "area_afetada", "label": "Área afetada", "type": "text"},
            {"name": "responsavel_id", "label": "Responsável pelo atendimento", "type": "choice", "choice": "pessoas"},
            {"name": "data_abertura", "label": "Data de identificação", "type": "date", "required": True},
            {"name": "data_encerramento", "label": "Data de encerramento", "type": "date"},
            {"name": "status", "label": "Status", "type": "select", "options": INCIDENT_STATUSES, "required": True},
            {"name": "causa", "label": "Causa provável", "type": "textarea", "span": 2},
            {"name": "acoes_tomadas", "label": "Ações tomadas", "type": "textarea", "span": 2},
            {"name": "licoes_aprendidas", "label": "Lições aprendidas", "type": "textarea", "span": 2},
        ],
    },
    "acoes": {
        "title": "Plano de ação",
        "table": "acoes",
        "module": "acoes",
        "list_query": """
            SELECT a.*, p.nome AS responsavel_nome, c.titulo AS controle_titulo,
                   r.titulo AS risco_titulo, i.numero AS incidente_numero
            FROM acoes a
            LEFT JOIN pessoas p ON p.id = a.responsavel_id
            LEFT JOIN controles c ON c.id = a.controle_id
            LEFT JOIN riscos r ON r.id = a.risco_id
            LEFT JOIN incidentes i ON i.id = a.incidente_id
            WHERE a.ativo = 1
        """,
        "id_column": "a.id",
        "search_fields": ["a.titulo", "a.descricao", "a.origem", "p.nome", "c.titulo", "r.titulo", "i.numero"],
        "filters": {"status": "a.status", "responsavel_id": "a.responsavel_id"},
        "columns": [
            ("titulo", "Ação"),
            ("origem", "Origem"),
            ("prioridade", "Prioridade"),
            ("responsavel_nome", "Responsável"),
            ("prazo", "Prazo"),
            ("status", "Status"),
        ],
        "fields": [
            {"name": "titulo", "label": "Título da ação", "type": "text", "required": True, "span": 2},
            {"name": "descricao", "label": "Descrição", "type": "textarea", "span": 2},
            {"name": "origem", "label": "Origem", "type": "select", "options": ["Controle", "Risco", "Incidente", "Auditoria", "Melhoria", "Outro"]},
            {"name": "prioridade", "label": "Prioridade", "type": "select", "options": ["Baixa", "Média", "Alta", "Crítica"], "required": True},
            {"name": "status", "label": "Status", "type": "select", "options": ACTION_STATUSES, "required": True},
            {"name": "responsavel_id", "label": "Responsável", "type": "choice", "choice": "pessoas"},
            {"name": "controle_id", "label": "Controle relacionado", "type": "choice", "choice": "controles"},
            {"name": "risco_id", "label": "Risco relacionado", "type": "choice", "choice": "riscos"},
            {"name": "incidente_id", "label": "Incidente relacionado", "type": "choice", "choice": "incidentes"},
            {"name": "data_inicio", "label": "Data inicial", "type": "date"},
            {"name": "prazo", "label": "Prazo final", "type": "date", "required": True},
            {"name": "data_conclusao", "label": "Data de conclusão", "type": "date"},
            {"name": "evidencia_conclusao", "label": "Evidência de conclusão", "type": "textarea", "span": 2},
            {"name": "observacoes", "label": "Observações", "type": "textarea", "span": 2},
        ],
    },
    "usuarios": {
        "title": "Usuários",
        "table": "usuarios",
        "module": "usuarios",
        "list_query": """
            SELECT u.id, u.nome, u.email, u.setor, u.status, u.criado_em, p.nome AS perfil_nome
            FROM usuarios u
            LEFT JOIN perfis p ON p.id = u.perfil_id
            WHERE 1=1
        """,
        "id_column": "u.id",
        "search_fields": ["u.nome", "u.email", "u.setor", "p.nome"],
        "filters": {"status": "u.status", "setor": "u.setor", "perfil_id": "u.perfil_id"},
        "columns": [
            ("nome", "Nome"),
            ("email", "E-mail"),
            ("perfil_nome", "Perfil"),
            ("setor", "Setor"),
            ("status", "Status"),
        ],
        "fields": [
            {"name": "nome", "label": "Nome", "type": "text", "required": True},
            {"name": "email", "label": "E-mail", "type": "email", "required": True},
            {"name": "senha", "label": "Senha", "type": "password"},
            {"name": "perfil_id", "label": "Perfil", "type": "choice", "choice": "perfis", "required": True},
            {"name": "setor", "label": "Setor", "type": "text"},
            {"name": "status", "label": "Status", "type": "select", "options": ["Ativo", "Inativo"], "required": True},
        ],
    },
    "perfis": {
        "title": "Perfis",
        "table": "perfis",
        "module": "perfis",
        "list_query": "SELECT p.* FROM perfis p WHERE 1=1",
        "id_column": "p.id",
        "search_fields": ["p.nome", "p.descricao"],
        "filters": {"status": "p.status"},
        "columns": [
            ("nome", "Nome"),
            ("descricao", "Descrição"),
            ("status", "Status"),
        ],
        "fields": [
            {"name": "nome", "label": "Nome", "type": "text", "required": True},
            {"name": "descricao", "label": "Descrição", "type": "textarea"},
            {"name": "status", "label": "Status", "type": "select", "options": ["Ativo", "Inativo"], "required": True},
            {"name": "permissoes", "label": "Permissões", "type": "permissions", "span": 2},
        ],
    },
}


def _get_admin_password() -> str:
    pw = os.environ.get("PGSI_ADMIN_PASSWORD", "")
    if not pw:
        pw = secrets.token_urlsafe(16)
        print(
            f"WARNING: PGSI_ADMIN_PASSWORD not set.\n"
            f"         Generated admin password: {pw}\n"
            f"         Set PGSI_ADMIN_PASSWORD env var or change this password immediately."
        )
    return pw


_login_attempts: dict[str, list[float]] = {}
_RATE_LIMIT_WINDOW = 60.0
_RATE_LIMIT_MAX = 10


def _check_rate_limit(ip: str) -> bool:
    now = time.monotonic()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < _RATE_LIMIT_WINDOW]
    _login_attempts[ip] = attempts
    if len(attempts) >= _RATE_LIMIT_MAX:
        return False
    _login_attempts[ip].append(now)
    return True


def _is_safe_url(url: str | None) -> bool:
    if not url:
        return False
    from urllib.parse import urlparse
    parsed = urlparse(url)
    return not parsed.netloc and not parsed.scheme


def _file_path(file_row: sqlite3.Row) -> Path:
    caminho = file_row["caminho"]
    p = Path(caminho)
    if p.is_absolute():
        return p
    return UPLOAD_DIR / caminho


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_: Exception | None = None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = get_db()
    db.executescript(SCHEMA_SQL)
    run_migrations(db)
    seed_database(db)
    db.commit()


def run_migrations(db: sqlite3.Connection) -> None:
    ensure_column(db, "controles", "eixo", "TEXT")
    ensure_column(db, "riscos", "eixo", "TEXT")
    ensure_column(db, "acoes", "eixo", "TEXT")
    ensure_column(db, "arquivos", "descricao", "TEXT")
    ensure_column(db, "arquivos", "eixo", "TEXT")
    backfill_axes(db)


def ensure_column(db: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    columns = {row["name"] for row in db.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


@app.cli.command("init-db")
def init_db_command() -> None:
    init_db()
    print(f"Banco inicializado em {DATABASE}")


def seed_database(db: sqlite3.Connection) -> None:
    roles = [
        ("Administrador", "Acesso total ao sistema"),
        ("Gestor", "Gestão de controles, riscos, indicadores e relatórios"),
        ("Técnico", "Atualização técnica, evidências, incidentes e controles"),
        ("Auditor", "Consulta, evidências, relatórios e logs"),
        ("Consulta", "Consulta ampla sem edição"),
        ("Usuário comum", "Envio de documentos e consulta limitada"),
    ]
    for role, description in roles:
        db.execute(
            """
            INSERT OR IGNORE INTO perfis (nome, descricao, permissoes, status)
            VALUES (?, ?, ?, 'Ativo')
            """,
            (role, description, json.dumps(permissions_for_role(role), ensure_ascii=False)),
        )
        row = db.execute("SELECT id, permissoes FROM perfis WHERE nome = ?", (role,)).fetchone()
        if row:
            try:
                current_permissions = json.loads(row["permissoes"] or "{}")
            except json.JSONDecodeError:
                current_permissions = {}
            merged_permissions = current_permissions | {
                module: sorted(set(current_permissions.get(module, []) + actions))
                for module, actions in permissions_for_role(role).items()
            }
            db.execute(
                "UPDATE perfis SET permissoes = ?, atualizado_em = CURRENT_TIMESTAMP WHERE id = ?",
                (json.dumps(merged_permissions, ensure_ascii=False), row["id"]),
            )

    admin_profile = db.execute("SELECT id FROM perfis WHERE nome = 'Administrador'").fetchone()["id"]
    admin_exists = db.execute("SELECT id FROM usuarios WHERE email = 'admin@pgsi.local'").fetchone()
    if not admin_exists:
        db.execute(
            """
            INSERT INTO usuarios (nome, email, senha_hash, perfil_id, setor, status)
            VALUES (?, ?, ?, ?, ?, 'Ativo')
            """,
            ("Administrador PGSI", "admin@pgsi.local", generate_password_hash(_get_admin_password()), admin_profile, "Segurança da Informação"),
        )

    if db.execute("SELECT COUNT(*) AS total FROM pessoas").fetchone()["total"] == 0:
        pessoas = [
            ("Ana Martins", "1001", "ana.martins@ifto.local", "(63) 99999-1001", "Gestora de Segurança da Informação", "Segurança da Informação", "Gestor", "Reitoria"),
            ("Bruno Silva", "1002", "bruno.silva@ifto.local", "(63) 99999-1002", "Analista de Infraestrutura", "Infraestrutura", "Técnico", "DTI"),
            ("Carla Nunes", "1003", "carla.nunes@ifto.local", "(63) 99999-1003", "Auditora", "Controle Interno", "Auditor", "Reitoria"),
            ("Diego Costa", "1004", "diego.costa@ifto.local", "(63) 99999-1004", "Analista de Redes", "Redes", "Técnico", "DTI"),
            ("Elisa Rocha", "1005", "elisa.rocha@ifto.local", "(63) 99999-1005", "Encarregada de Dados", "Privacidade", "Gestor", "Reitoria"),
        ]
        db.executemany(
            """
            INSERT INTO pessoas
                (nome, matricula, email, telefone, cargo, setor, perfil_acesso, unidade_administrativa, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Ativo')
            """,
            pessoas,
        )

    pessoa_ids = [row["id"] for row in db.execute("SELECT id FROM pessoas ORDER BY id").fetchall()]
    if db.execute("SELECT COUNT(*) AS total FROM controles").fetchone()["total"] == 0:
        statuses = ["Implementado", "Em andamento", "Pendente", "Parcialmente implementado", "Em revisão", "Vencido"]
        base_date = TODAY - timedelta(days=220)
        for index, item in enumerate(PGSI_CONTROLS):
            titulo, categoria, tipo, descricao, area, periodicidade, criticidade = item
            status = statuses[index % len(statuses)]
            created = base_date + timedelta(days=index * 4)
            revision = TODAY + timedelta(days=(index % 8 - 3) * 35)
            deadline = TODAY + timedelta(days=(index % 7 - 2) * 20)
            db.execute(
                """
                INSERT INTO controles
                    (titulo, descricao, categoria, tipo, area_responsavel, responsavel_id,
                     eixo, periodicidade, status, criticidade, data_criacao, data_atualizacao,
                     data_revisao, prazo, observacoes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    titulo,
                    descricao,
                    categoria,
                    tipo,
                    area,
                    pessoa_ids[index % len(pessoa_ids)],
                    area_title_for_control(titulo, categoria),
                    periodicidade,
                    status,
                    criticidade,
                    created.isoformat(),
                    (created + timedelta(days=30)).isoformat(),
                    revision.isoformat(),
                    deadline.isoformat(),
                    "Controle inicial baseado nos temas do PGSI.",
                ),
            )
    ensure_security_area_controls(db, pessoa_ids)

    if db.execute("SELECT COUNT(*) AS total FROM riscos").fetchone()["total"] == 0:
        sample_risks = [
            ("Ausência de plano de recuperação testado", "Continuidade", "Ambiente de backup", 3, 3, 11, "Em tratamento", "Executar teste trimestral de restauração e registrar evidências.", 30),
            ("Contas privilegiadas sem revisão periódica", "Controle de Acesso", "Diretório institucional", 2, 3, 5, "Monitorado", "Revisar permissões administrativas mensalmente.", 45),
            ("Vulnerabilidades críticas sem priorização", "Vulnerabilidades", "Servidores expostos", 3, 2, 7, "Identificado", "", 15),
            ("Baixa adesão a treinamentos", "Conscientização", "Usuários finais", 2, 2, 14, "Em tratamento", "Campanha anual de conscientização e trilha por perfil.", 90),
        ]
        for title, category, asset, probability, impact, control_id, status, treatment, days in sample_risks:
            score, level = calculate_risk(probability, impact)
            db.execute(
                """
                INSERT INTO riscos
                    (titulo, descricao, categoria, eixo, ativo_relacionado, controle_id, probabilidade,
                     impacto, pontuacao, nivel_risco, responsavel_id, plano_tratamento, prazo, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    title,
                    f"Risco relacionado a {category.lower()} identificado no acompanhamento inicial.",
                    category,
                    area_title_for_control(title, category),
                    asset,
                    control_id,
                    probability,
                    impact,
                    score,
                    level,
                    pessoa_ids[0],
                    treatment,
                    (TODAY + timedelta(days=days)).isoformat(),
                    status,
                ),
            )

    if db.execute("SELECT COUNT(*) AS total FROM incidentes").fetchone()["total"] == 0:
        incidents = [
            ("INC-2026-001", "Tentativa de phishing reportada", "Phishing", "Média", "E-mail institucional", "DTI", "Em andamento", TODAY - timedelta(days=8), None),
            ("INC-2026-002", "Falha de backup em servidor", "Falha de backup", "Alta", "Servidor de arquivos", "Infraestrutura", "Aberto", TODAY - timedelta(days=21), None),
            ("INC-2026-003", "Indisponibilidade de sistema acadêmico", "Indisponibilidade de sistema", "Média", "Sistema acadêmico", "Sistemas", "Encerrado", TODAY - timedelta(days=40), TODAY - timedelta(days=39)),
        ]
        for number, title, kind, severity, asset, area, status, opened, closed in incidents:
            db.execute(
                """
                INSERT INTO incidentes
                    (numero, titulo, descricao, tipo, gravidade, ativo_afetado, area_afetada,
                     status, responsavel_id, data_abertura, data_encerramento, causa, acoes_tomadas)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    number,
                    title,
                    "Registro inicial para acompanhamento do ciclo de resposta a incidentes.",
                    kind,
                    severity,
                    asset,
                    area,
                    status,
                    pessoa_ids[1],
                    opened.isoformat(),
                    closed.isoformat() if closed else None,
                    "Em análise" if not closed else "Falha operacional identificada",
                    "Triagem realizada e ações em andamento." if not closed else "Serviço restabelecido e relatório concluído.",
                ),
            )

    if db.execute("SELECT COUNT(*) AS total FROM acoes").fetchone()["total"] == 0:
        actions = [
            ("Revisar controles vencidos", "Controle", "Alta", "Em andamento", pessoa_ids[0], 6, None, None, TODAY - timedelta(days=20), TODAY + timedelta(days=10)),
            ("Solicitar evidências de backup", "Controle", "Crítica", "Aberta", pessoa_ids[1], 11, 1, None, TODAY - timedelta(days=15), TODAY - timedelta(days=2)),
            ("Formalizar comunicação de incidentes", "Incidente", "Alta", "Aguardando validação", pessoa_ids[0], 17, None, 2, TODAY - timedelta(days=30), TODAY + timedelta(days=5)),
            ("Atualizar inventário de fornecedores", "Melhoria", "Média", "Aberta", pessoa_ids[2], 15, None, None, TODAY - timedelta(days=10), TODAY + timedelta(days=60)),
        ]
        for title, origin, priority, status, owner, control_id, risk_id, incident_id, start, deadline in actions:
            db.execute(
                """
                INSERT INTO acoes
                    (titulo, descricao, origem, prioridade, status, responsavel_id,
                     controle_id, risco_id, incidente_id, data_inicio, prazo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    title,
                    "Ação corretiva/preventiva gerada a partir dos pontos de atenção iniciais.",
                    origin,
                    priority,
                    status,
                    owner,
                    control_id,
                    risk_id,
                    incident_id,
                    start.isoformat(),
                    deadline.isoformat(),
                ),
            )

    if db.execute("SELECT COUNT(*) AS total FROM indicadores").fetchone()["total"] == 0:
        indicators = [
            ("Percentual de controles implementados", "Avaliar avanço da implantação", "controles implementados / total de controles x 100", "Mensal"),
            ("Percentual de riscos altos", "Priorizar riscos críticos", "riscos altos / total de riscos x 100", "Mensal"),
            ("Ações vencidas", "Apontar falhas de execução", "ações com prazo menor que hoje e status diferente de concluída", "Semanal"),
            ("Incidentes em aberto", "Medir volume de resposta ativa", "incidentes com status aberto ou em andamento", "Semanal"),
            ("Controles sem evidência", "Identificar controles sem documentação", "controles que não possuem arquivo associado", "Mensal"),
        ]
        db.executemany(
            """
            INSERT INTO indicadores (nome, descricao, formula, periodicidade, referencia)
            VALUES (?, ?, ?, ?, 'PGSI')
            """,
            indicators,
        )


def current_date() -> date:
    return date.today()


@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; style-src 'self' 'unsafe-inline' https://use.typekit.net; "
        "script-src 'self' 'unsafe-inline' https://use.typekit.net; "
        "font-src 'self' data: https://use.typekit.net https://p.typekit.net; "
        "img-src 'self' data: https://use.typekit.net",
    )
    return response


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def format_date(value: Any) -> str:
    if not value:
        return ""
    parsed = parse_date(str(value))
    if not parsed:
        return str(value)
    return parsed.strftime("%d/%m/%Y")


def calculate_risk(probability: Any, impact: Any) -> tuple[int, str]:
    probability_int = int(probability or 0)
    impact_int = int(impact or 0)
    score = probability_int * impact_int
    if score <= 2:
        return score, "Baixo"
    if score <= 5:
        return score, "Médio"
    return score, "Alto"


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any]:
    return dict(row) if row else {}


def audit(action: str, table: str | None = None, record_id: int | None = None, details: str | None = None) -> None:
    db = get_db()
    user_id = session.get("user_id")
    db.execute(
        """
        INSERT INTO logs_auditoria (usuario_id, acao, tabela_afetada, registro_id, ip, detalhes)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, action, table, record_id, request.remote_addr, details),
    )
    db.commit()


def load_permissions(profile_row: sqlite3.Row | None) -> dict[str, list[str]]:
    if not profile_row:
        return {}
    try:
        return json.loads(profile_row["permissoes"] or "{}")
    except json.JSONDecodeError:
        return {}


@app.before_request
def load_logged_user() -> None:
    db = get_db()
    user_id = session.get("user_id")
    g.user = None
    g.profile = None
    g.permissions = {}
    if user_id:
        user = db.execute(
            """
            SELECT u.*, p.nome AS perfil_nome, p.permissoes
            FROM usuarios u
            LEFT JOIN perfis p ON p.id = u.perfil_id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()
        if user and user["status"] == "Ativo":
            g.user = user
            g.permissions = load_permissions(user)
        else:
            session.clear()


def has_permission(module: str, action: str = "view") -> bool:
    if not g.get("user"):
        return False
    if g.user["perfil_nome"] == "Administrador":
        return True
    return action in g.permissions.get(module, [])


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not g.get("user"):
            return redirect(url_for("login", next=request.full_path))
        return view(*args, **kwargs)

    return wrapped


def permission_required(module: str, action: str = "view"):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not g.get("user"):
                return redirect(url_for("login", next=request.full_path))
            if not has_permission(module, action):
                abort(403)
            return view(*args, **kwargs)

        return wrapped

    return decorator


@app.context_processor
def inject_globals() -> dict[str, Any]:
    nav_sections = [
        {
            "title": "Principal",
            "items": [
                {"module": "dashboard", "label": "Dashboard", "endpoint": "dashboard", "params": {}, "hint": "Indicadores gerais"},
                {"module": "relatorios", "label": "Relatórios", "endpoint": "reports", "params": {}, "hint": "Exportações e visão executiva"},
            ],
        },
        {
            "title": "Gestão de controles",
            "items": [
                {
                    "module": "controles",
                    "label": "Controles",
                    "endpoint": "control_center",
                    "params": {},
                    "hint": "Controles, riscos, ações, evidências e melhorias",
                    "area": "controles",
                },
            ],
        },
        {
            "title": "Operação",
            "items": [
                {"module": "incidentes", "label": "Incidentes", "endpoint": "entity_list", "params": {"entity": "incidentes"}, "hint": "Registro e resposta"},
                {"module": "pessoas", "label": "Pessoas", "endpoint": "entity_list", "params": {"entity": "pessoas"}, "hint": "Responsáveis e áreas"},
            ],
        },
        {
            "title": "Acesso e sistema",
            "items": [
                {"module": "usuarios", "label": "Usuários", "endpoint": "entity_list", "params": {"entity": "usuarios"}, "hint": "Contas do sistema"},
                {"module": "perfis", "label": "Perfis", "endpoint": "entity_list", "params": {"entity": "perfis"}, "hint": "Permissões por módulo"},
                {"module": "logs", "label": "Auditoria", "endpoint": "logs", "params": {}, "hint": "Histórico de alterações"},
                {"module": "backup", "label": "Backup", "endpoint": "backup", "params": {}, "hint": "Cópias do banco"},
            ],
        },
    ]
    return {
        "app_version": APP_VERSION,
        "current_user": g.get("user"),
        "has_permission": has_permission,
        "nav_sections": nav_sections,
        "control_tabs": build_control_tabs(),
        "is_nav_active": is_nav_active,
        "module_labels": MODULE_LABELS,
        "action_labels": ACTION_LABELS,
        "permission_matrix": PERMISSION_MATRIX,
        "format_date": format_date,
        "today_iso": current_date().isoformat(),
    }


def build_control_tabs() -> list[dict[str, Any]]:
    return [
        {
            "key": "visao",
            "label": "Visão geral",
            "description": "Indicadores, alertas e atalhos",
            "module": "controles",
            "action": "view",
            "endpoint": "control_center",
            "params": {},
        },
        {
            "key": "controles",
            "label": "Controles",
            "description": "Cadastro, status, revisões e responsáveis",
            "module": "controles",
            "action": "view",
            "endpoint": "entity_list",
            "params": {"entity": "controles"},
        },
        {
            "key": "riscos",
            "label": "Riscos",
            "description": "Probabilidade, impacto e tratamento",
            "module": "riscos",
            "action": "view",
            "endpoint": "entity_list",
            "params": {"entity": "riscos"},
        },
        {
            "key": "acoes",
            "label": "Plano de ação",
            "description": "Prazos, prioridades e execução",
            "module": "acoes",
            "action": "view",
            "endpoint": "entity_list",
            "params": {"entity": "acoes"},
        },
        {
            "key": "arquivos",
            "label": "Evidências",
            "description": "Uploads, documentos e extração",
            "module": "arquivos",
            "action": "view",
            "endpoint": "documents",
            "params": {},
        },
        {
            "key": "melhorias",
            "label": "Melhorias",
            "description": "Pontos de atenção e recomendações",
            "module": "relatorios",
            "action": "view",
            "endpoint": "reports",
            "params": {"tipo": "melhorias"},
        },
    ]


def control_area_for_entity(entity: str | None) -> str | None:
    if entity in {"controles", "riscos", "acoes"}:
        return entity
    return None


def is_nav_active(item: dict[str, Any]) -> bool:
    endpoint = request.endpoint or ""
    if item.get("area") == "controles":
        if endpoint in {"security_areas", "security_area_detail"}:
            return True
        if endpoint == "control_center":
            return True
        if endpoint in {"documents", "file_detail", "file_download", "file_view", "file_analyze", "file_delete"}:
            return True
        if endpoint in {"entity_list", "entity_create", "entity_edit", "entity_delete"}:
            return request.view_args and request.view_args.get("entity") in {"controles", "riscos", "acoes"}
        if endpoint in {"reports", "report_export"}:
            return request.args.get("tipo") in {"controles", "riscos", "melhorias", "evidencias"}
        return False
    if endpoint != item["endpoint"]:
        return False
    params = item.get("params") or {}
    return all(request.view_args and request.view_args.get(key) == value for key, value in params.items()) if params else True


@app.template_filter("status_class")
def status_class(value: Any) -> str:
    normalized = str(value or "").lower()
    if any(word in normalized for word in ["venc", "alto", "alta", "crítica", "critica", "crítico", "critico", "aberto", "descart", "desativ"]):
        return "danger"
    if any(word in normalized for word in ["pendente", "andamento", "revisão", "revisao", "médio", "medio", "manutenção", "manutencao", "reservado", "aguardando"]):
        return "warning"
    if any(word in normalized for word in ["implementado", "concluído", "concluido", "encerrado", "baixo", "baixa", "ativo", "mitigado"]):
        return "success"
    return "neutral"


def get_choice_options(kind: str) -> list[tuple[str, str]]:
    db = get_db()
    if kind == "pessoas":
        rows = db.execute("SELECT id, nome FROM pessoas WHERE status = 'Ativo' ORDER BY nome").fetchall()
        return [(str(row["id"]), row["nome"]) for row in rows]
    if kind == "perfis":
        rows = db.execute("SELECT id, nome FROM perfis WHERE status = 'Ativo' ORDER BY nome").fetchall()
        return [(str(row["id"]), row["nome"]) for row in rows]
    if kind == "controles":
        rows = db.execute("SELECT id, titulo FROM controles WHERE ativo = 1 ORDER BY titulo").fetchall()
        return [(str(row["id"]), row["titulo"]) for row in rows]
    if kind == "riscos":
        rows = db.execute("SELECT id, titulo FROM riscos WHERE ativo = 1 ORDER BY titulo").fetchall()
        return [(str(row["id"]), row["titulo"]) for row in rows]
    if kind == "incidentes":
        rows = db.execute("SELECT id, numero || ' - ' || titulo AS nome FROM incidentes WHERE ativo = 1 ORDER BY data_abertura DESC").fetchall()
        return [(str(row["id"]), row["nome"]) for row in rows]
    if kind == "acoes":
        rows = db.execute("SELECT id, titulo FROM acoes WHERE ativo = 1 ORDER BY titulo").fetchall()
        return [(str(row["id"]), row["titulo"]) for row in rows]
    return []


def build_select_choices(config: dict[str, Any]) -> dict[str, list[tuple[str, str]]]:
    choices = {}
    for field in config["fields"]:
        if field.get("type") == "choice":
            choices[field["name"]] = get_choice_options(field["choice"])
    return choices


def normalize_field_value(field: dict[str, Any], form: dict[str, str]) -> Any:
    value = form.get(field["name"], "").strip()
    if field.get("type") == "choice":
        return int(value) if value else None
    if field.get("type") == "select" and field.get("options"):
        return value
    if field.get("type") == "date":
        return value or None
    return value or None


def collect_entity_payload(entity: str, item_id: int | None = None) -> tuple[dict[str, Any], list[str]]:
    config = ENTITY_CONFIGS[entity]
    payload: dict[str, Any] = {}
    errors: list[str] = []
    for field in config["fields"]:
        name = field["name"]
        if field["type"] == "permissions":
            payload[name] = json.dumps(collect_permissions_from_form(), ensure_ascii=False)
            continue
        if entity == "usuarios" and name == "senha":
            password = request.form.get("senha", "")
            if item_id is None and not password:
                errors.append("A senha é obrigatória para novo usuário.")
            if password:
                if len(password) < 8:
                    errors.append("A senha deve ter pelo menos 8 caracteres.")
                else:
                    payload["senha_hash"] = generate_password_hash(password)
            continue
        value = normalize_field_value(field, request.form)
        if field.get("required") and value in (None, ""):
            errors.append(f"{field['label']} é obrigatório.")
        payload[name] = value

    if entity == "controles":
        if not payload.get("responsavel_id"):
            errors.append("Todo controle deve possuir um responsável.")
        payload["data_atualizacao"] = payload.get("data_atualizacao") or current_date().isoformat()
        payload["eixo"] = request.form.get("eixo") or area_title_for_control(payload.get("titulo"), payload.get("categoria"), payload.get("tipo"), payload.get("descricao"))

    if entity == "riscos":
        probability = payload.get("probabilidade") or 0
        impact = payload.get("impacto") or 0
        score, level = calculate_risk(probability, impact)
        payload["probabilidade"] = int(probability)
        payload["impacto"] = int(impact)
        payload["pontuacao"] = score
        payload["nivel_risco"] = level
        payload["eixo"] = request.form.get("eixo") or area_title_for_control(payload.get("titulo"), payload.get("categoria"), payload.get("descricao"), payload.get("ativo_relacionado"))
        if level == "Alto" and not payload.get("plano_tratamento"):
            errors.append("Todo risco alto deve possuir plano de tratamento.")

    if entity == "incidentes" and not payload.get("status"):
        errors.append("Todo incidente deve possuir status.")

    if entity == "acoes" and not payload.get("prazo"):
        errors.append("Toda ação deve possuir prazo.")
    if entity == "acoes":
        payload["eixo"] = request.form.get("eixo") or area_title_for_control(payload.get("titulo"), payload.get("origem"), payload.get("descricao"))

    if entity == "perfis" and not payload.get("permissoes"):
        payload["permissoes"] = "{}"

    return payload, errors


def collect_permissions_from_form() -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for module, actions in PERMISSION_MATRIX.items():
        selected = []
        for action in actions:
            if request.form.get(f"perm__{module}__{action}") == "on":
                selected.append(action)
        if selected:
            result[module] = selected
    return result


def insert_entity(entity: str, payload: dict[str, Any]) -> int:
    table = ENTITY_CONFIGS[entity]["table"]
    columns = list(payload.keys())
    placeholders = ", ".join("?" for _ in columns)
    sql = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})"
    cursor = get_db().execute(sql, [payload[column] for column in columns])
    get_db().commit()
    return int(cursor.lastrowid)


def update_entity(entity: str, item_id: int, payload: dict[str, Any]) -> None:
    table = ENTITY_CONFIGS[entity]["table"]
    payload["atualizado_em"] = datetime.utcnow().isoformat(timespec="seconds") if table in {"pessoas", "usuarios", "perfis", "riscos", "incidentes", "acoes"} else payload.get("atualizado_em")
    payload = {key: value for key, value in payload.items() if value is not None or key != "atualizado_em"}
    assignments = ", ".join(f"{column} = ?" for column in payload.keys())
    get_db().execute(f"UPDATE {table} SET {assignments} WHERE id = ?", [*payload.values(), item_id])
    get_db().commit()


def fetch_entity(entity: str, item_id: int) -> sqlite3.Row | None:
    table = ENTITY_CONFIGS[entity]["table"]
    return get_db().execute(f"SELECT * FROM {table} WHERE id = ?", (item_id,)).fetchone()


def prepare_rows(rows: list[sqlite3.Row], columns: list[tuple[str, str]]) -> list[dict[str, Any]]:
    prepared = []
    date_fields = {"data_criacao", "data_atualizacao", "data_revisao", "prazo", "data_abertura", "data_encerramento", "data_inicio", "data_conclusao", "criado_em", "atualizado_em"}
    for row in rows:
        item = dict(row)
        for key, _ in columns:
            if key in date_fields:
                item[key] = format_date(item.get(key))
            elif key in {"probabilidade", "impacto"}:
                item[key] = {1: "Baixa", 2: "Média", 3: "Alta"}.get(item.get(key), item.get(key))
            elif key == "evidencias":
                item[key] = str(item.get(key) or 0)
        prepared.append(item)
    return prepared


def query_entity_list(entity: str) -> list[sqlite3.Row]:
    config = ENTITY_CONFIGS[entity]
    sql = config["list_query"]
    params: list[Any] = []
    query = request.args.get("q", "").strip()
    if query:
        like = f"%{query}%"
        search_sql = " OR ".join(f"{field} LIKE ?" for field in config["search_fields"])
        sql += f" AND ({search_sql})"
        params.extend([like] * len(config["search_fields"]))
    for filter_name, column in config.get("filters", {}).items():
        value = request.args.get(filter_name, "").strip()
        if value:
            sql += f" AND {column} = ?"
            params.append(value)
    sql += f" ORDER BY {config['id_column']} DESC"
    return get_db().execute(sql, params).fetchall()


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.get("user"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        client_ip = request.remote_addr or "unknown"
        if not _check_rate_limit(client_ip):
            flash("Muitas tentativas. Aguarde 1 minuto e tente novamente.", "error")
            return render_template("login.html")
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("senha", "")
        user = get_db().execute(
            """
            SELECT u.*, p.nome AS perfil_nome, p.permissoes
            FROM usuarios u
            LEFT JOIN perfis p ON p.id = u.perfil_id
            WHERE lower(u.email) = ?
            """,
            (email,),
        ).fetchone()
        if user and user["status"] == "Ativo" and check_password_hash(user["senha_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session.permanent = True
            audit("login", "usuarios", user["id"], "Login realizado")
            next_url = request.args.get("next")
            return redirect(next_url if _is_safe_url(next_url) else url_for("dashboard"))
        flash("E-mail ou senha inválidos.", "error")
    return render_template("login.html")


@app.route("/logout", methods=["POST"])
@login_required
def logout():
    audit("logout", "usuarios", g.user["id"], "Logout realizado")
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    if not g.get("user"):
        return redirect(url_for("login"))
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
@permission_required("dashboard", "view")
def dashboard():
    data = dashboard_data()
    filters = {
        "inicio": request.args.get("inicio", ""),
        "fim": request.args.get("fim", ""),
        "setor": request.args.get("setor", ""),
        "responsavel_id": request.args.get("responsavel_id", ""),
        "categoria": request.args.get("categoria", ""),
    }
    return render_template(
        "dashboard.html",
        data=data,
        filters=filters,
        people=get_choice_options("pessoas"),
    )


@app.route("/controles/painel")
@permission_required("controles", "view")
def control_center():
    return render_template(
        "control_center.html",
        data=control_center_data(),
        control_area="visao",
    )


def control_center_data() -> dict[str, Any]:
    db = get_db()
    today = current_date().isoformat()
    total_controls = db.execute("SELECT COUNT(*) AS total FROM controles WHERE ativo = 1").fetchone()["total"]
    implemented = db.execute("SELECT COUNT(*) AS total FROM controles WHERE ativo = 1 AND status IN ('Implementado', 'Concluído')").fetchone()["total"]
    pending = db.execute("SELECT COUNT(*) AS total FROM controles WHERE ativo = 1 AND status IN ('Pendente', 'Não iniciado', 'Parcialmente implementado', 'Em revisão')").fetchone()["total"]
    evidence_gap = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM controles c
        WHERE c.ativo = 1
          AND NOT EXISTS (
              SELECT 1 FROM arquivos a WHERE a.controle_id = c.id AND a.ativo = 1
          )
        """
    ).fetchone()["total"]
    overdue_actions = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM acoes
        WHERE ativo = 1 AND prazo < ? AND status NOT IN ('Concluída', 'Cancelada')
        """,
        (today,),
    ).fetchone()["total"]
    high_risks = db.execute("SELECT COUNT(*) AS total FROM riscos WHERE ativo = 1 AND nivel_risco = 'Alto'").fetchone()["total"]
    evidence_count = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM arquivos
        WHERE ativo = 1 AND (controle_id IS NOT NULL OR risco_id IS NOT NULL OR acao_id IS NOT NULL)
        """
    ).fetchone()["total"]

    recent_controls = db.execute(
        """
        SELECT c.id, c.titulo, c.categoria, c.status, c.prazo, p.nome AS responsavel_nome
        FROM controles c
        LEFT JOIN pessoas p ON p.id = c.responsavel_id
        WHERE c.ativo = 1
        ORDER BY COALESCE(c.data_atualizacao, c.data_criacao) DESC, c.id DESC
        LIMIT 8
        """
    ).fetchall()
    critical_risks = db.execute(
        """
        SELECT r.id, r.titulo, r.nivel_risco, r.pontuacao, r.prazo, c.titulo AS controle_titulo
        FROM riscos r
        LEFT JOIN controles c ON c.id = r.controle_id
        WHERE r.ativo = 1
        ORDER BY r.pontuacao DESC, r.prazo
        LIMIT 6
        """
    ).fetchall()
    next_actions = db.execute(
        """
        SELECT a.id, a.titulo, a.prioridade, a.status, a.prazo, p.nome AS responsavel_nome
        FROM acoes a
        LEFT JOIN pessoas p ON p.id = a.responsavel_id
        WHERE a.ativo = 1 AND a.status NOT IN ('Concluída', 'Cancelada')
        ORDER BY a.prazo
        LIMIT 6
        """
    ).fetchall()
    evidence_rows = db.execute(
        """
        SELECT a.id, a.nome_original, a.classificacao, a.criado_em, c.titulo AS controle_titulo,
               r.titulo AS risco_titulo, ac.titulo AS acao_titulo
        FROM arquivos a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN acoes ac ON ac.id = a.acao_id
        WHERE a.ativo = 1 AND (a.controle_id IS NOT NULL OR a.risco_id IS NOT NULL OR a.acao_id IS NOT NULL)
        ORDER BY a.criado_em DESC
        LIMIT 6
        """
    ).fetchall()
    alerts, improvements = generate_alerts_and_improvements(db)
    implementation_pct = round((implemented / total_controls) * 100, 1) if total_controls else 0
    return {
        "cards": [
            {"label": "Controles", "value": total_controls, "hint": f"{implementation_pct}% implementados", "tone": "neutral"},
            {"label": "Pendências", "value": pending, "hint": "Controles não concluídos", "tone": "warning"},
            {"label": "Sem evidência", "value": evidence_gap, "hint": "Controles a documentar", "tone": "warning"},
            {"label": "Riscos altos", "value": high_risks, "hint": "Prioridade de tratamento", "tone": "danger"},
            {"label": "Ações vencidas", "value": overdue_actions, "hint": "Prazo expirado", "tone": "danger"},
            {"label": "Evidências", "value": evidence_count, "hint": "Documentos vinculados", "tone": "success"},
        ],
        "recent_controls": recent_controls,
        "critical_risks": critical_risks,
        "next_actions": next_actions,
        "evidence_rows": evidence_rows,
        "alerts": alerts[:5],
        "improvements": improvements[:5],
        "areas": security_area_overview(),
    }


@app.route("/areas")
@permission_required("controles", "view")
def security_areas():
    return redirect(url_for("control_center") + "#areas-pgsi")


@app.route("/areas/<key>")
@permission_required("controles", "view")
def security_area_detail(key: str):
    area = get_security_area(key)
    if not area:
        abort(404)
    tabs = ["visao", "alimentacao", "controles", "riscos", "acoes", "evidencias", "indicadores", "historico"]
    if area["key"] == "ativos":
        tabs.insert(2, "inventario")
    active_tab = request.args.get("tab", "visao")
    if active_tab not in tabs:
        active_tab = "visao"
    return render_template(
        "area_detail.html",
        area=area,
        data=security_area_detail_data(area),
        active_tab=active_tab,
        area_tabs=tabs,
        control_area="visao",
    )


@app.route("/areas/ativos/inventario", methods=["POST"])
@permission_required("controles", "create")
def asset_inventory_create():
    errors = create_asset_inventory_item(request.form)
    if errors:
        for error in errors:
            flash(error, "error")
    else:
        flash("Item adicionado ao inventário de ativos.", "success")
    return redirect(url_for("security_area_detail", key="ativos", tab="inventario"))


@app.route("/areas/ativos/inventario/<int:item_id>/inativar", methods=["POST"])
@permission_required("controles", "delete")
def asset_inventory_delete(item_id: int):
    item = get_db().execute("SELECT id, nome FROM ativos_inventario WHERE id = ? AND ativo = 1", (item_id,)).fetchone()
    if not item:
        abort(404)
    get_db().execute("UPDATE ativos_inventario SET ativo = 0, atualizado_em = CURRENT_TIMESTAMP WHERE id = ?", (item_id,))
    get_db().commit()
    audit("inativar", "ativos_inventario", item_id, f"Item de inventário inativado: {item['nome']}")
    flash("Item removido do inventário ativo.", "success")
    return redirect(url_for("security_area_detail", key="ativos", tab="inventario"))


def get_security_area(key: str) -> dict[str, Any] | None:
    return next((area for area in SECURITY_AREAS if area["key"] == key), None)


def area_title_for_control(*values: Any) -> str | None:
    haystack = " ".join(str(value or "").lower() for value in values)
    for area in SECURITY_AREAS:
        if area["title"].lower() in haystack:
            return area["title"]
        if any(term.lower() in haystack for term in area["terms"]):
            return area["title"]
    return None


def backfill_axes(db: sqlite3.Connection) -> None:
    for row in db.execute("SELECT id, titulo, descricao, categoria, tipo, area_responsavel FROM controles WHERE eixo IS NULL OR trim(eixo) = ''").fetchall():
        eixo = area_title_for_control(row["titulo"], row["descricao"], row["categoria"], row["tipo"], row["area_responsavel"])
        if eixo:
            db.execute("UPDATE controles SET eixo = ? WHERE id = ?", (eixo, row["id"]))
    for row in db.execute(
        """
        SELECT r.id, r.titulo, r.descricao, r.categoria, r.ativo_relacionado, c.eixo AS controle_eixo, c.titulo AS controle_titulo
        FROM riscos r
        LEFT JOIN controles c ON c.id = r.controle_id
        WHERE r.eixo IS NULL OR trim(r.eixo) = ''
        """
    ).fetchall():
        eixo = row["controle_eixo"] or area_title_for_control(row["titulo"], row["descricao"], row["categoria"], row["ativo_relacionado"], row["controle_titulo"])
        if eixo:
            db.execute("UPDATE riscos SET eixo = ? WHERE id = ?", (eixo, row["id"]))
    for row in db.execute(
        """
        SELECT a.id, a.titulo, a.descricao, a.origem, c.eixo AS controle_eixo, r.eixo AS risco_eixo, i.titulo AS incidente_titulo
        FROM acoes a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        WHERE a.eixo IS NULL OR trim(a.eixo) = ''
        """
    ).fetchall():
        eixo = row["controle_eixo"] or row["risco_eixo"] or area_title_for_control(row["titulo"], row["descricao"], row["origem"], row["incidente_titulo"])
        if eixo:
            db.execute("UPDATE acoes SET eixo = ? WHERE id = ?", (eixo, row["id"]))
    for row in db.execute(
        """
        SELECT a.id, a.nome_original, a.descricao, a.classificacao, c.eixo AS controle_eixo,
               r.eixo AS risco_eixo, ac.eixo AS acao_eixo
        FROM arquivos a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN acoes ac ON ac.id = a.acao_id
        WHERE a.eixo IS NULL OR trim(a.eixo) = ''
        """
    ).fetchall():
        eixo = row["controle_eixo"] or row["risco_eixo"] or row["acao_eixo"] or area_title_for_control(row["nome_original"], row["descricao"], row["classificacao"])
        if eixo:
            db.execute("UPDATE arquivos SET eixo = ? WHERE id = ?", (eixo, row["id"]))


def asset_inventory_rows() -> list[sqlite3.Row]:
    return get_db().execute(
        """
        SELECT a.*, p.nome AS responsavel_nome
        FROM ativos_inventario a
        LEFT JOIN pessoas p ON p.id = a.responsavel_id
        WHERE a.ativo = 1
        ORDER BY
            CASE a.criticidade
                WHEN 'Crítica' THEN 1
                WHEN 'Alta' THEN 2
                WHEN 'Média' THEN 3
                WHEN 'Baixa' THEN 4
                ELSE 5
            END,
            a.nome
        """
    ).fetchall()


def asset_inventory_metrics() -> dict[str, int]:
    today = current_date().isoformat()
    total = scalar("SELECT COUNT(*) AS total FROM ativos_inventario WHERE ativo = 1")
    critical = scalar(
        "SELECT COUNT(*) AS total FROM ativos_inventario WHERE ativo = 1 AND criticidade IN ('Alta', 'Crítica')"
    )
    without_owner = scalar(
        "SELECT COUNT(*) AS total FROM ativos_inventario WHERE ativo = 1 AND responsavel_id IS NULL"
    )
    review_due = scalar(
        """
        SELECT COUNT(*) AS total
        FROM ativos_inventario
        WHERE ativo = 1
          AND data_revisao IS NOT NULL
          AND data_revisao < ?
        """,
        (today,),
    )
    maintenance = scalar(
        "SELECT COUNT(*) AS total FROM ativos_inventario WHERE ativo = 1 AND status = 'Em manutenção'"
    )
    return {
        "total": total,
        "critical": critical,
        "without_owner": without_owner,
        "review_due": review_due,
        "maintenance": maintenance,
    }


def create_asset_inventory_item(form: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    nome = (form.get("nome") or "").strip()
    if not nome:
        errors.append("Informe o nome do item do inventário.")
    if errors:
        return errors

    responsavel = (form.get("responsavel_id") or "").strip()
    try:
        responsavel_id = int(responsavel) if responsavel else None
    except ValueError:
        return ["Responsável inválido para o item do inventário."]
    criticidade = (form.get("criticidade") or "").strip() or "Média"
    status = (form.get("status") or "").strip() or "Ativo"
    if criticidade not in ASSET_CRITICALITIES:
        return ["Criticidade inválida para o item do inventário."]
    if status not in ASSET_STATUSES:
        return ["Status inválido para o item do inventário."]
    cursor = get_db().execute(
        """
        INSERT INTO ativos_inventario
            (nome, identificador, tipo, responsavel_id, criticidade, status,
             localizacao, unidade, data_aquisicao, data_revisao, observacoes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            nome,
            (form.get("identificador") or "").strip() or None,
            (form.get("tipo") or "").strip() or None,
            responsavel_id,
            criticidade,
            status,
            (form.get("localizacao") or "").strip() or None,
            (form.get("unidade") or "").strip() or None,
            (form.get("data_aquisicao") or "").strip() or None,
            (form.get("data_revisao") or "").strip() or None,
            (form.get("observacoes") or "").strip() or None,
        ),
    )
    get_db().commit()
    audit("criar", "ativos_inventario", cursor.lastrowid, f"Item de inventário criado: {nome}")
    return []


def area_playbook(area: dict[str, Any]) -> dict[str, list[str]]:
    return AREA_PLAYBOOKS.get(
        area["key"],
        {
            "records": ["Controle relacionado", "Risco identificado", "Plano de ação", "Evidência documental"],
            "evidence": ["Documento comprobatório", "Relatório técnico", "Registro de revisão", "Plano aprovado"],
            "workflow": ["Cadastrar controle", "Avaliar riscos", "Criar ações", "Anexar evidências"],
            "questions": ["Qual informação precisa ser registrada?", "Quem é o responsável?", "Qual prazo de revisão?"],
        },
    )


def area_clause(fields: list[str], area: dict[str, Any]) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    for term in area["terms"]:
        for field in fields:
            clauses.append(f"LOWER(COALESCE({field}, '')) LIKE LOWER(?)")
            params.append(f"%{term}%")
    return f"({' OR '.join(clauses)})", params


def scalar(sql: str, params: list[Any] | tuple[Any, ...] = ()) -> int:
    row = get_db().execute(sql, params).fetchone()
    return int(row["total"] if row else 0)


def ensure_security_area_controls(db: sqlite3.Connection, pessoa_ids: list[int]) -> None:
    if not pessoa_ids:
        return
    statuses = ["Em andamento", "Pendente", "Em revisão"]
    for index, area in enumerate(SECURITY_AREAS):
        control_filter, params = area_clause(["c.eixo", "c.titulo", "c.descricao", "c.categoria", "c.tipo", "c.area_responsavel"], area)
        existing = db.execute(
            f"SELECT COUNT(*) AS total FROM controles c WHERE c.ativo = 1 AND {control_filter}",
            params,
        ).fetchone()["total"]
        if existing:
            continue
        db.execute(
            """
            INSERT INTO controles
                (titulo, descricao, categoria, tipo, area_responsavel, responsavel_id,
                 eixo, periodicidade, status, criticidade, data_criacao, data_atualizacao,
                 data_revisao, prazo, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                f"Programa - {area['title']}",
                area["explanation"],
                area["title"],
                area["control_type"],
                area["responsible_area"],
                pessoa_ids[index % len(pessoa_ids)],
                area["title"],
                "Anual",
                statuses[index % len(statuses)],
                "Alta" if index % 3 else "Crítica",
                current_date().isoformat(),
                current_date().isoformat(),
                (current_date() + timedelta(days=180)).isoformat(),
                (current_date() + timedelta(days=90)).isoformat(),
                f"Controle-base criado para estruturar a área PGSI: {area['title']}.",
            ),
        )


def security_area_metrics(area: dict[str, Any]) -> dict[str, Any]:
    db = get_db()
    today = current_date().isoformat()
    control_filter, control_params = area_clause(["c.eixo", "c.titulo", "c.descricao", "c.categoria", "c.tipo", "c.area_responsavel"], area)
    risk_filter, risk_params = area_clause(["r.eixo", "r.titulo", "r.descricao", "r.categoria", "r.ativo_relacionado", "c.eixo", "c.titulo", "c.categoria"], area)
    action_filter, action_params = area_clause(["a.eixo", "a.titulo", "a.descricao", "a.origem", "c.eixo", "c.titulo", "r.eixo", "r.titulo", "i.titulo", "i.tipo"], area)
    file_filter, file_params = area_clause(["a.eixo", "a.nome_original", "a.classificacao", "a.descricao", "c.eixo", "c.titulo", "c.categoria", "r.eixo", "r.titulo", "r.categoria", "ac.eixo", "ac.titulo"], area)
    inventory = asset_inventory_metrics() if area["key"] == "ativos" else {"total": 0, "critical": 0, "without_owner": 0, "review_due": 0, "maintenance": 0}

    controls = scalar(f"SELECT COUNT(*) AS total FROM controles c WHERE c.ativo = 1 AND {control_filter}", control_params)
    implemented_controls = scalar(
        f"SELECT COUNT(*) AS total FROM controles c WHERE c.ativo = 1 AND c.status IN ('Implementado', 'Concluído') AND {control_filter}",
        control_params,
    )
    controls_without_evidence = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM controles c
        WHERE c.ativo = 1
          AND {control_filter}
          AND NOT EXISTS (SELECT 1 FROM arquivos a WHERE a.controle_id = c.id AND a.ativo = 1)
        """,
        control_params,
    )
    risks = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM riscos r
        LEFT JOIN controles c ON c.id = r.controle_id
        WHERE r.ativo = 1 AND {risk_filter}
        """,
        risk_params,
    )
    high_risks = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM riscos r
        LEFT JOIN controles c ON c.id = r.controle_id
        WHERE r.ativo = 1 AND r.nivel_risco = 'Alto' AND {risk_filter}
        """,
        risk_params,
    )
    high_risks_without_plan = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM riscos r
        LEFT JOIN controles c ON c.id = r.controle_id
        WHERE r.ativo = 1
          AND r.nivel_risco = 'Alto'
          AND (r.plano_tratamento IS NULL OR trim(r.plano_tratamento) = '')
          AND {risk_filter}
        """,
        risk_params,
    )
    actions = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM acoes a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        WHERE a.ativo = 1 AND {action_filter}
        """,
        action_params,
    )
    overdue_actions = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM acoes a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        WHERE a.ativo = 1
          AND a.prazo < ?
          AND a.status NOT IN ('Concluída', 'Cancelada')
          AND {action_filter}
        """,
        [today, *action_params],
    )
    open_actions = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM acoes a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        WHERE a.ativo = 1
          AND a.status NOT IN ('Concluída', 'Cancelada')
          AND {action_filter}
        """,
        action_params,
    )
    completed_actions = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM acoes a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        WHERE a.ativo = 1
          AND a.status = 'Concluída'
          AND {action_filter}
        """,
        action_params,
    )
    evidences = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM arquivos a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN acoes ac ON ac.id = a.acao_id
        WHERE a.ativo = 1 AND {file_filter}
        """,
        file_params,
    )
    overdue_controls = scalar(
        f"""
        SELECT COUNT(*) AS total
        FROM controles c
        WHERE c.ativo = 1
          AND c.prazo IS NOT NULL
          AND c.prazo < ?
          AND c.status NOT IN ('Implementado', 'Concluído')
          AND {control_filter}
        """,
        [today, *control_params],
    )
    attention = overdue_controls + high_risks_without_plan + overdue_actions + controls_without_evidence + inventory["without_owner"] + inventory["review_due"]
    implemented_pct = round((implemented_controls / controls) * 100, 1) if controls else 0
    completed_actions_pct = round((completed_actions / actions) * 100, 1) if actions else 0
    status, tone = area_status(
        controls=controls,
        evidences=evidences,
        high_risks=high_risks,
        high_risks_without_plan=high_risks_without_plan,
        overdue_actions=overdue_actions,
        overdue_controls=overdue_controls,
        controls_without_evidence=controls_without_evidence,
    )
    if area["key"] == "ativos" and not inventory["total"]:
        status, tone = "Inventário pendente", "warning"
    elif area["key"] == "ativos" and (inventory["without_owner"] or inventory["review_due"]):
        status, tone = "Em atenção", "warning"
    return {
        "controls": controls,
        "implemented_controls": implemented_controls,
        "implemented_pct": implemented_pct,
        "controls_without_evidence": controls_without_evidence,
        "overdue_controls": overdue_controls,
        "risks": risks,
        "high_risks": high_risks,
        "high_risks_without_plan": high_risks_without_plan,
        "actions": actions,
        "open_actions": open_actions,
        "completed_actions": completed_actions,
        "completed_actions_pct": completed_actions_pct,
        "overdue_actions": overdue_actions,
        "evidences": evidences,
        "attention": attention,
        "status": status,
        "tone": tone,
        "inventory_total": inventory["total"],
        "inventory_critical": inventory["critical"],
        "inventory_without_owner": inventory["without_owner"],
        "inventory_review_due": inventory["review_due"],
        "inventory_maintenance": inventory["maintenance"],
    }


def area_status(**metrics: int) -> tuple[str, str]:
    if metrics["controls"] == 0 and metrics["evidences"] == 0:
        return "Sem informações suficientes", "neutral"
    if metrics["overdue_actions"] or metrics["overdue_controls"] or metrics["high_risks_without_plan"]:
        return "Crítico", "danger"
    if metrics["high_risks"] or metrics["controls_without_evidence"]:
        return "Em atenção", "warning"
    if metrics["evidences"] and metrics["controls"]:
        return "Regular", "success"
    return "Em evolução", "warning"


def security_area_overview() -> list[dict[str, Any]]:
    overview = []
    for area in SECURITY_AREAS:
        overview.append({**area, "metrics": security_area_metrics(area)})
    return overview


def security_area_detail_data(area: dict[str, Any]) -> dict[str, Any]:
    db = get_db()
    control_filter, control_params = area_clause(["c.eixo", "c.titulo", "c.descricao", "c.categoria", "c.tipo", "c.area_responsavel"], area)
    risk_filter, risk_params = area_clause(["r.eixo", "r.titulo", "r.descricao", "r.categoria", "r.ativo_relacionado", "c.eixo", "c.titulo", "c.categoria"], area)
    action_filter, action_params = area_clause(["a.eixo", "a.titulo", "a.descricao", "a.origem", "c.eixo", "c.titulo", "r.eixo", "r.titulo", "i.titulo", "i.tipo"], area)
    file_filter, file_params = area_clause(["a.eixo", "a.nome_original", "a.classificacao", "a.descricao", "c.eixo", "c.titulo", "c.categoria", "r.eixo", "r.titulo", "r.categoria", "ac.eixo", "ac.titulo"], area)
    controls = db.execute(
        f"""
        SELECT c.id, c.titulo, c.descricao, c.categoria, c.tipo, c.eixo, c.status, c.criticidade,
               c.periodicidade, c.data_criacao, c.data_atualizacao, c.data_revisao, c.prazo,
               p.nome AS responsavel_nome,
               (SELECT COUNT(*) FROM arquivos a WHERE a.controle_id = c.id AND a.ativo = 1) AS evidencias
        FROM controles c
        LEFT JOIN pessoas p ON p.id = c.responsavel_id
        WHERE c.ativo = 1 AND {control_filter}
        ORDER BY c.criticidade DESC, c.prazo
        LIMIT 30
        """,
        control_params,
    ).fetchall()
    risks = db.execute(
        f"""
        SELECT r.id, r.titulo, r.descricao, r.categoria, r.eixo, r.nivel_risco, r.pontuacao,
               r.probabilidade, r.impacto, r.status, r.plano_tratamento, r.prazo,
               p.nome AS responsavel_nome, c.titulo AS controle_titulo
        FROM riscos r
        LEFT JOIN controles c ON c.id = r.controle_id
        LEFT JOIN pessoas p ON p.id = r.responsavel_id
        WHERE r.ativo = 1 AND {risk_filter}
        ORDER BY r.pontuacao DESC, r.prazo
        LIMIT 20
        """,
        risk_params,
    ).fetchall()
    actions = db.execute(
        f"""
        SELECT a.id, a.titulo, a.descricao, a.origem, a.eixo, a.prioridade, a.status,
               a.data_inicio, a.prazo, a.data_conclusao, p.nome AS responsavel_nome
        FROM acoes a
        LEFT JOIN pessoas p ON p.id = a.responsavel_id
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        WHERE a.ativo = 1 AND {action_filter}
        ORDER BY a.prazo
        LIMIT 20
        """,
        action_params,
    ).fetchall()
    evidences = db.execute(
        f"""
        SELECT a.id, a.nome_original, a.classificacao, a.descricao, a.eixo, a.criado_em, c.titulo AS controle_titulo,
               r.titulo AS risco_titulo, ac.titulo AS acao_titulo
        FROM arquivos a
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN acoes ac ON ac.id = a.acao_id
        WHERE a.ativo = 1 AND {file_filter}
        ORDER BY a.criado_em DESC
        LIMIT 20
        """,
        file_params,
    ).fetchall()
    inventory = asset_inventory_rows() if area["key"] == "ativos" else []
    inventory_metrics = asset_inventory_metrics() if area["key"] == "ativos" else {"total": 0, "critical": 0, "without_owner": 0, "review_due": 0, "maintenance": 0}
    metrics = security_area_metrics(area)
    history = security_area_history(controls, risks, actions, evidences, inventory)
    last_update = latest_area_update(controls, risks, actions, evidences, inventory)
    responsible = main_area_responsible(controls, risks, actions, inventory)
    attention_points = security_area_attention_points(area, controls, risks, actions, metrics, inventory)
    return {
        "metrics": metrics,
        "status": {"label": metrics["status"], "tone": metrics["tone"]},
        "last_update": last_update,
        "responsible": responsible,
        "playbook": area_playbook(area),
        "attention_points": attention_points,
        "controls": controls,
        "risks": risks,
        "actions": actions,
        "evidences": evidences,
        "inventory": inventory,
        "inventory_metrics": inventory_metrics,
        "asset_options": {
            "types": ASSET_TYPES,
            "criticalities": ASSET_CRITICALITIES,
            "statuses": ASSET_STATUSES,
        },
        "people": get_choice_options("pessoas"),
        "choices": {
            "controles": [(str(row["id"]), row["titulo"]) for row in controls],
            "riscos": [(str(row["id"]), row["titulo"]) for row in risks],
            "acoes": [(str(row["id"]), row["titulo"]) for row in actions],
        },
        "history": history,
    }


def latest_area_update(
    controls: list[sqlite3.Row],
    risks: list[sqlite3.Row],
    actions: list[sqlite3.Row],
    evidences: list[sqlite3.Row],
    inventory: list[sqlite3.Row] | None = None,
) -> str:
    candidates: list[date] = []
    for row in controls:
        for field in ("data_atualizacao", "data_criacao"):
            parsed = parse_date(row[field])
            if parsed:
                candidates.append(parsed)
    for row in risks:
        parsed = parse_date(row["prazo"])
        if parsed:
            candidates.append(parsed)
    for row in actions:
        for field in ("data_conclusao", "data_inicio"):
            parsed = parse_date(row[field])
            if parsed:
                candidates.append(parsed)
    for row in evidences:
        parsed = parse_date(row["criado_em"])
        if parsed:
            candidates.append(parsed)
    for row in inventory or []:
        for field in ("data_revisao", "data_aquisicao", "atualizado_em", "criado_em"):
            parsed = parse_date(row[field])
            if parsed:
                candidates.append(parsed)
    return max(candidates).isoformat() if candidates else ""


def main_area_responsible(
    controls: list[sqlite3.Row],
    risks: list[sqlite3.Row],
    actions: list[sqlite3.Row],
    inventory: list[sqlite3.Row] | None = None,
) -> str:
    counts: dict[str, int] = {}
    for collection in (controls, risks, actions, inventory or []):
        for row in collection:
            name = row["responsavel_nome"] if "responsavel_nome" in row.keys() else None
            if name:
                counts[name] = counts.get(name, 0) + 1
    if not counts:
        return "Sem responsável definido"
    return sorted(counts.items(), key=lambda item: item[1], reverse=True)[0][0]


def security_area_attention_points(
    area: dict[str, Any],
    controls: list[sqlite3.Row],
    risks: list[sqlite3.Row],
    actions: list[sqlite3.Row],
    metrics: dict[str, Any],
    inventory: list[sqlite3.Row] | None = None,
) -> list[dict[str, str]]:
    today = current_date()
    points: list[dict[str, str]] = []
    if area["key"] == "ativos" and not inventory:
        points.append({"level": "warning", "title": "Inventário vazio", "text": "Cadastre os itens que compõem o inventário de ativos deste ambiente."})
    if metrics["evidences"] == 0:
        points.append({"level": "warning", "title": "Eixo sem evidências", "text": "Anexe documentos para comprovar a execução dos controles deste eixo."})
    for control in controls:
        deadline = parse_date(control["prazo"])
        if deadline and deadline < today and control["status"] not in {"Implementado", "Concluído"}:
            points.append({"level": "danger", "title": "Controle vencido", "text": control["titulo"]})
        if int(control["evidencias"] or 0) == 0:
            points.append({"level": "warning", "title": "Controle sem evidência", "text": control["titulo"]})
        if control["status"] == "Parcialmente implementado":
            points.append({"level": "warning", "title": "Controle parcialmente implementado", "text": control["titulo"]})
    for risk in risks:
        if risk["nivel_risco"] == "Alto" and not (risk["plano_tratamento"] or "").strip():
            points.append({"level": "danger", "title": "Risco alto sem plano", "text": risk["titulo"]})
    for action in actions:
        deadline = parse_date(action["prazo"])
        if deadline and deadline < today and action["status"] not in {"Concluída", "Cancelada"}:
            points.append({"level": "danger", "title": "Ação vencida", "text": action["titulo"]})
    for item in inventory or []:
        if not item["responsavel_nome"]:
            points.append({"level": "warning", "title": "Ativo sem responsável", "text": item["nome"]})
        review_date = parse_date(item["data_revisao"])
        if review_date and review_date < today:
            points.append({"level": "danger", "title": "Revisão de ativo vencida", "text": item["nome"]})
    if not points:
        points.append({"level": "success", "title": "Sem alerta crítico", "text": f"{area['title']} não possui pontos críticos automáticos no momento."})
    return points[:12]


def security_area_history(
    controls: list[sqlite3.Row],
    risks: list[sqlite3.Row],
    actions: list[sqlite3.Row],
    evidences: list[sqlite3.Row],
    inventory: list[sqlite3.Row] | None = None,
) -> list[sqlite3.Row]:
    references = {
        "controles": [row["id"] for row in controls],
        "riscos": [row["id"] for row in risks],
        "acoes": [row["id"] for row in actions],
        "arquivos": [row["id"] for row in evidences],
        "ativos_inventario": [row["id"] for row in inventory or []],
    }
    clauses = []
    params: list[Any] = []
    for table, ids in references.items():
        if not ids:
            continue
        placeholders = ", ".join("?" for _ in ids)
        clauses.append(f"(l.tabela_afetada = ? AND l.registro_id IN ({placeholders}))")
        params.extend([table, *ids])
    if not clauses:
        return []
    return get_db().execute(
        f"""
        SELECT l.*, u.nome AS usuario_nome
        FROM logs_auditoria l
        LEFT JOIN usuarios u ON u.id = l.usuario_id
        WHERE {' OR '.join(clauses)}
        ORDER BY l.id DESC
        LIMIT 40
        """,
        params,
    ).fetchall()


def dashboard_data() -> dict[str, Any]:
    db = get_db()
    today = current_date().isoformat()
    control_where = ["c.ativo = 1"]
    params: list[Any] = []
    if request.args.get("inicio"):
        control_where.append("c.data_criacao >= ?")
        params.append(request.args["inicio"])
    if request.args.get("fim"):
        control_where.append("c.data_criacao <= ?")
        params.append(request.args["fim"])
    if request.args.get("setor"):
        control_where.append("c.area_responsavel = ?")
        params.append(request.args["setor"])
    if request.args.get("responsavel_id"):
        control_where.append("c.responsavel_id = ?")
        params.append(request.args["responsavel_id"])
    if request.args.get("categoria"):
        control_where.append("c.categoria = ?")
        params.append(request.args["categoria"])
    where_sql = " AND ".join(control_where)

    total_controls = db.execute(f"SELECT COUNT(*) AS total FROM controles c WHERE {where_sql}", params).fetchone()["total"]
    implemented = db.execute(
        f"SELECT COUNT(*) AS total FROM controles c WHERE {where_sql} AND c.status IN ('Implementado', 'Concluído')",
        params,
    ).fetchone()["total"]
    pending = db.execute(
        f"SELECT COUNT(*) AS total FROM controles c WHERE {where_sql} AND c.status IN ('Pendente', 'Não iniciado', 'Parcialmente implementado', 'Em revisão')",
        params,
    ).fetchone()["total"]
    overdue_controls = db.execute(
        f"""
        SELECT COUNT(*) AS total FROM controles c
        WHERE {where_sql}
          AND (c.status = 'Vencido' OR (c.prazo IS NOT NULL AND c.prazo < ? AND c.status NOT IN ('Implementado', 'Concluído')))
        """,
        [*params, today],
    ).fetchone()["total"]
    high_risks = db.execute("SELECT COUNT(*) AS total FROM riscos WHERE ativo = 1 AND nivel_risco = 'Alto'").fetchone()["total"]
    open_incidents = db.execute("SELECT COUNT(*) AS total FROM incidentes WHERE ativo = 1 AND status IN ('Aberto', 'Em andamento', 'Contido')").fetchone()["total"]
    overdue_actions = db.execute(
        """
        SELECT COUNT(*) AS total FROM acoes
        WHERE ativo = 1 AND prazo < ? AND status NOT IN ('Concluída', 'Cancelada')
        """,
        (today,),
    ).fetchone()["total"]
    documents = db.execute("SELECT COUNT(*) AS total FROM arquivos WHERE ativo = 1").fetchone()["total"]

    status_rows = db.execute(
        f"SELECT c.status, COUNT(*) AS total FROM controles c WHERE {where_sql} GROUP BY c.status ORDER BY total DESC",
        params,
    ).fetchall()
    risk_rows = db.execute("SELECT nivel_risco, COUNT(*) AS total FROM riscos WHERE ativo = 1 GROUP BY nivel_risco").fetchall()
    monthly_rows = db.execute(
        """
        SELECT substr(data_abertura, 1, 7) AS mes, COUNT(*) AS total
        FROM incidentes
        WHERE ativo = 1
        GROUP BY substr(data_abertura, 1, 7)
        ORDER BY mes
        """
    ).fetchall()
    recent_logs = db.execute(
        """
        SELECT l.*, u.nome AS usuario_nome
        FROM logs_auditoria l
        LEFT JOIN usuarios u ON u.id = l.usuario_id
        ORDER BY l.id DESC
        LIMIT 8
        """
    ).fetchall()

    implementation_pct = round((implemented / total_controls) * 100, 1) if total_controls else 0
    risk_total = db.execute("SELECT COUNT(*) AS total FROM riscos WHERE ativo = 1").fetchone()["total"]
    high_risk_pct = round((high_risks / risk_total) * 100, 1) if risk_total else 0

    alerts, improvements = generate_alerts_and_improvements(db)
    return {
        "cards": [
            {"label": "Controles cadastrados", "value": total_controls, "tone": "neutral"},
            {"label": "Controles implementados", "value": implemented, "hint": f"{implementation_pct}% do total", "tone": "success"},
            {"label": "Controles pendentes", "value": pending, "tone": "warning"},
            {"label": "Controles vencidos", "value": overdue_controls, "tone": "danger"},
            {"label": "Riscos altos", "value": high_risks, "hint": f"{high_risk_pct}% dos riscos", "tone": "danger"},
            {"label": "Incidentes abertos", "value": open_incidents, "tone": "warning"},
            {"label": "Ações vencidas", "value": overdue_actions, "tone": "danger"},
            {"label": "Documentos enviados", "value": documents, "tone": "neutral"},
        ],
        "charts": {
            "status": [{"label": row["status"], "value": row["total"]} for row in status_rows],
            "risks": [{"label": row["nivel_risco"], "value": row["total"]} for row in risk_rows],
            "incidents": [{"label": row["mes"], "value": row["total"]} for row in monthly_rows],
        },
        "alerts": alerts,
        "improvements": improvements,
        "recent_logs": recent_logs,
    }


def generate_alerts_and_improvements(db: sqlite3.Connection) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    today = current_date()
    alerts: list[dict[str, str]] = []
    improvements: list[dict[str, str]] = []

    controls_without_evidence = db.execute(
        """
        SELECT c.id, c.titulo
        FROM controles c
        LEFT JOIN arquivos a ON a.controle_id = c.id AND a.ativo = 1
        WHERE c.ativo = 1
        GROUP BY c.id
        HAVING COUNT(a.id) = 0
        LIMIT 8
        """
    ).fetchall()
    for row in controls_without_evidence:
        alerts.append({"level": "warning", "title": "Controle sem evidência", "text": row["titulo"]})
    if controls_without_evidence:
        improvements.append({"title": "Solicitar evidências", "text": "Priorize controles implementados ou críticos sem documentação vinculada."})

    expired_controls = db.execute(
        """
        SELECT id, titulo, prazo
        FROM controles
        WHERE ativo = 1 AND prazo IS NOT NULL AND prazo < ? AND status NOT IN ('Implementado', 'Concluído')
        LIMIT 8
        """,
        (today.isoformat(),),
    ).fetchall()
    for row in expired_controls:
        alerts.append({"level": "danger", "title": "Controle vencido", "text": f"{row['titulo']} venceu em {format_date(row['prazo'])}"})
    if expired_controls:
        improvements.append({"title": "Revisar controles vencidos", "text": "Atualize prazos, status e responsáveis dos controles atrasados."})

    high_risks_without_plan = db.execute(
        """
        SELECT id, titulo
        FROM riscos
        WHERE ativo = 1 AND nivel_risco = 'Alto' AND (plano_tratamento IS NULL OR trim(plano_tratamento) = '')
        LIMIT 8
        """
    ).fetchall()
    for row in high_risks_without_plan:
        alerts.append({"level": "danger", "title": "Risco alto sem plano", "text": row["titulo"]})
    if high_risks_without_plan:
        improvements.append({"title": "Priorizar riscos altos", "text": "Defina tratamento, prazo e responsável para os riscos classificados como altos."})

    old_incidents = db.execute(
        """
        SELECT numero, titulo, data_abertura
        FROM incidentes
        WHERE ativo = 1 AND status IN ('Aberto', 'Em andamento', 'Contido')
        LIMIT 20
        """
    ).fetchall()
    for row in old_incidents:
        opened = parse_date(row["data_abertura"])
        if opened and (today - opened).days > 15:
            alerts.append({"level": "warning", "title": "Incidente aberto há muitos dias", "text": f"{row['numero']} - {row['titulo']}"})
    if any((parse_date(row["data_abertura"]) and (today - parse_date(row["data_abertura"])).days > 15) for row in old_incidents):
        improvements.append({"title": "Acelerar resposta a incidentes", "text": "Revise causa provável, contenção e comunicação dos incidentes antigos."})

    overdue_actions = db.execute(
        """
        SELECT titulo, prazo
        FROM acoes
        WHERE ativo = 1 AND prazo < ? AND status NOT IN ('Concluída', 'Cancelada')
        LIMIT 8
        """,
        (today.isoformat(),),
    ).fetchall()
    for row in overdue_actions:
        alerts.append({"level": "danger", "title": "Ação vencida", "text": f"{row['titulo']} venceu em {format_date(row['prazo'])}"})
    if overdue_actions:
        improvements.append({"title": "Replanejar ações vencidas", "text": "Negocie novo prazo ou registre conclusão com evidência."})

    annual_review = db.execute(
        """
        SELECT titulo, data_revisao
        FROM controles
        WHERE ativo = 1 AND data_revisao IS NOT NULL AND data_revisao < ?
        LIMIT 8
        """,
        (today.isoformat(),),
    ).fetchall()
    for row in annual_review:
        alerts.append({"level": "warning", "title": "Revisão anual pendente", "text": row["titulo"]})
    if annual_review:
        improvements.append({"title": "Atualizar documentação de conformidade", "text": "Registre revisão anual dos controles e evidências vinculadas."})

    if not alerts:
        alerts.append({"level": "success", "title": "Sem pontos críticos", "text": "Nenhum alerta automático foi identificado no momento."})
    if not improvements:
        improvements.append({"title": "Manter rotina de monitoramento", "text": "Continue acompanhando indicadores, prazos e evidências."})

    return alerts[:12], improvements[:8]


@app.route("/<entity>")
@login_required
def entity_list(entity: str):
    if entity not in ENTITY_CONFIGS:
        abort(404)
    config = ENTITY_CONFIGS[entity]
    if not has_permission(config["module"], "view"):
        abort(403)
    rows = query_entity_list(entity)
    return render_template(
        "entity_list.html",
        entity=entity,
        config=config,
        rows=prepare_rows(rows, config["columns"]),
        filters=request.args,
        people=get_choice_options("pessoas"),
        profiles=get_choice_options("perfis"),
        control_area=control_area_for_entity(entity),
    )


@app.route("/<entity>/novo", methods=["GET", "POST"])
@login_required
def entity_create(entity: str):
    if entity not in ENTITY_CONFIGS:
        abort(404)
    config = ENTITY_CONFIGS[entity]
    if not has_permission(config["module"], "create"):
        abort(403)
    values = {field["name"]: field.get("default", "") for field in config["fields"]}
    if entity == "controles":
        values["data_criacao"] = current_date().isoformat()
    if entity == "incidentes":
        values["data_abertura"] = current_date().isoformat()
    if entity == "acoes":
        values["data_inicio"] = current_date().isoformat()
    for field in config["fields"]:
        if field["name"] in request.args:
            values[field["name"]] = request.args.get(field["name"], "")
    errors: list[str] = []
    if request.method == "POST":
        payload, errors = collect_entity_payload(entity)
        values.update(request.form.to_dict())
        if not errors:
            try:
                item_id = insert_entity(entity, payload)
                audit("criar", config["table"], item_id, f"{config['title']} criado")
                flash(f"{config['title']} cadastrado com sucesso.", "success")
                return redirect(url_for("entity_list", entity=entity))
            except sqlite3.IntegrityError as exc:
                errors.append(f"Não foi possível salvar: {exc}")
    return render_template(
        "entity_form.html",
        entity=entity,
        config=config,
        item=None,
        values=values,
        errors=errors,
        choices=build_select_choices(config),
        permission_values={},
        control_area=control_area_for_entity(entity),
    )


@app.route("/<entity>/<int:item_id>/editar", methods=["GET", "POST"])
@login_required
def entity_edit(entity: str, item_id: int):
    if entity not in ENTITY_CONFIGS:
        abort(404)
    config = ENTITY_CONFIGS[entity]
    if not has_permission(config["module"], "edit"):
        abort(403)
    item = fetch_entity(entity, item_id)
    if not item:
        abort(404)
    values = row_to_dict(item)
    errors: list[str] = []
    if request.method == "POST":
        payload, errors = collect_entity_payload(entity, item_id)
        values.update(request.form.to_dict())
        if not errors:
            try:
                update_entity(entity, item_id, payload)
                audit("editar", config["table"], item_id, f"{config['title']} atualizado")
                flash(f"{config['title']} atualizado com sucesso.", "success")
                return redirect(url_for("entity_list", entity=entity))
            except sqlite3.IntegrityError as exc:
                errors.append(f"Não foi possível salvar: {exc}")
    permission_values = {}
    if entity == "perfis":
        try:
            permission_values = json.loads(item["permissoes"] or "{}")
        except json.JSONDecodeError:
            permission_values = {}
    return render_template(
        "entity_form.html",
        entity=entity,
        config=config,
        item=item,
        values=values,
        errors=errors,
        choices=build_select_choices(config),
        permission_values=permission_values,
        control_area=control_area_for_entity(entity),
    )


@app.route("/<entity>/<int:item_id>/inativar", methods=["POST"])
@login_required
def entity_delete(entity: str, item_id: int):
    if entity not in ENTITY_CONFIGS:
        abort(404)
    config = ENTITY_CONFIGS[entity]
    if not has_permission(config["module"], "delete"):
        abort(403)
    table = config["table"]
    if table in {"usuarios", "perfis", "pessoas"}:
        get_db().execute(f"UPDATE {table} SET status = 'Inativo', atualizado_em = CURRENT_TIMESTAMP WHERE id = ?", (item_id,))
    else:
        get_db().execute(f"UPDATE {table} SET ativo = 0, atualizado_em = CURRENT_TIMESTAMP WHERE id = ?", (item_id,))
    get_db().commit()
    audit("inativar", table, item_id, f"{config['title']} inativado")
    flash(f"Registro inativado em {config['title']}.", "success")
    return redirect(url_for("entity_list", entity=entity))


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def can_access_file(file_row: sqlite3.Row, action: str = "view") -> bool:
    if not g.get("user"):
        return False
    if g.user["perfil_nome"] == "Administrador":
        return True
    restricted = file_row["classificacao"] in {"Restrito", "Confidencial", "Sensível"}
    if restricted:
        privileged = g.user["perfil_nome"] in {"Gestor", "Auditor"} and has_permission("arquivos", action)
        return privileged or file_row["usuario_upload_id"] == g.user["id"]
    return has_permission("arquivos", action) or file_row["usuario_upload_id"] == g.user["id"]


@app.route("/arquivos", methods=["GET", "POST"])
@login_required
def documents():
    if request.method == "POST":
        if not has_permission("arquivos", "create"):
            abort(403)
        upload = request.files.get("arquivo")
        if not upload or not upload.filename:
            flash("Selecione um arquivo para envio.", "error")
            return redirect(url_for("documents"))
        if not allowed_file(upload.filename):
            flash("Formato de arquivo não autorizado.", "error")
            return redirect(url_for("documents"))
        original = secure_filename(upload.filename)
        ext = original.rsplit(".", 1)[1].lower()
        saved = f"{uuid.uuid4().hex}.{ext}"
        path = UPLOAD_DIR / saved
        upload.save(path)
        mime = mimetypes.guess_type(original)[0] or "application/octet-stream"
        classification = request.form.get("classificacao", "Interno")
        cursor = get_db().execute(
            """
            INSERT INTO arquivos
                (nome_original, nome_salvo, tipo, caminho, tamanho, classificacao,
                 descricao, eixo, usuario_upload_id, controle_id, risco_id, incidente_id, acao_id, pessoa_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                original,
                saved,
                mime,
                saved,
                path.stat().st_size,
                classification,
                request.form.get("descricao", "").strip() or None,
                request.form.get("eixo") or None,
                g.user["id"],
                request.form.get("controle_id") or None,
                request.form.get("risco_id") or None,
                request.form.get("incidente_id") or None,
                request.form.get("acao_id") or None,
                request.form.get("pessoa_id") or None,
            ),
        )
        get_db().commit()
        audit("upload", "arquivos", cursor.lastrowid, f"Arquivo enviado: {original}")
        flash("Documento enviado com sucesso.", "success")
        next_url = request.form.get("next")
        if _is_safe_url(next_url):
            return redirect(next_url)
        return redirect(url_for("file_detail", file_id=cursor.lastrowid))

    if not has_permission("arquivos", "view"):
        abort(403)
    rows = get_db().execute(
        """
        SELECT a.*, u.nome AS usuario_nome, c.titulo AS controle_titulo, r.titulo AS risco_titulo,
               i.numero AS incidente_numero, ac.titulo AS acao_titulo, p.nome AS pessoa_nome
        FROM arquivos a
        LEFT JOIN usuarios u ON u.id = a.usuario_upload_id
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        LEFT JOIN acoes ac ON ac.id = a.acao_id
        LEFT JOIN pessoas p ON p.id = a.pessoa_id
        WHERE a.ativo = 1
        ORDER BY a.id DESC
        """
    ).fetchall()
    visible_rows = [row for row in rows if can_access_file(row, "view")]
    return render_template(
        "documents.html",
        rows=visible_rows,
        choices={
            "controles": get_choice_options("controles"),
            "riscos": get_choice_options("riscos"),
            "incidentes": get_choice_options("incidentes"),
            "acoes": get_choice_options("acoes"),
            "pessoas": get_choice_options("pessoas"),
        },
        defaults=request.args,
        control_area="arquivos",
    )


@app.route("/arquivos/<int:file_id>")
@permission_required("arquivos", "view")
def file_detail(file_id: int):
    file_row = get_file(file_id)
    if not can_access_file(file_row, "view"):
        abort(403)
    return render_template("file_detail.html", file=file_row, control_area="arquivos")


def get_file(file_id: int) -> sqlite3.Row:
    file_row = get_db().execute(
        """
        SELECT a.*, u.nome AS usuario_nome
        FROM arquivos a
        LEFT JOIN usuarios u ON u.id = a.usuario_upload_id
        WHERE a.id = ? AND a.ativo = 1
        """,
        (file_id,),
    ).fetchone()
    if not file_row:
        abort(404)
    return file_row


@app.route("/arquivos/<int:file_id>/download")
@login_required
def file_download(file_id: int):
    file_row = get_file(file_id)
    if not can_access_file(file_row, "download"):
        abort(403)
    return send_file(_file_path(file_row), as_attachment=True, download_name=file_row["nome_original"])


@app.route("/arquivos/<int:file_id>/visualizar")
@login_required
def file_view(file_id: int):
    file_row = get_file(file_id)
    if not can_access_file(file_row, "view"):
        abort(403)
    return send_file(_file_path(file_row), as_attachment=False, download_name=file_row["nome_original"])


@app.route("/arquivos/<int:file_id>/analisar", methods=["POST"])
@login_required
def file_analyze(file_id: int):
    file_row = get_file(file_id)
    if not can_access_file(file_row, "analyze"):
        abort(403)
    text, imported = extract_file_text(file_row, import_controls=request.form.get("importar_controles") == "on")
    get_db().execute(
        "UPDATE arquivos SET extraido_texto = ?, extraido_em = CURRENT_TIMESTAMP WHERE id = ?",
        (text[:50000], file_id),
    )
    get_db().commit()
    audit("analisar", "arquivos", file_id, "Extração automática executada")
    if imported:
        flash(f"Análise concluída e {imported} controle(s) importado(s).", "success")
    else:
        flash("Análise concluída.", "success")
    return redirect(url_for("file_detail", file_id=file_id))


@app.route("/arquivos/<int:file_id>/inativar", methods=["POST"])
@permission_required("arquivos", "delete")
def file_delete(file_id: int):
    file_row = get_file(file_id)
    if not can_access_file(file_row, "delete"):
        abort(403)
    get_db().execute("UPDATE arquivos SET ativo = 0 WHERE id = ?", (file_id,))
    get_db().commit()
    audit("inativar", "arquivos", file_id, "Arquivo inativado")
    flash("Documento inativado.", "success")
    return redirect(url_for("documents"))


def extract_file_text(file_row: sqlite3.Row, import_controls: bool = False) -> tuple[str, int]:
    path = _file_path(file_row)
    ext = path.suffix.lower().lstrip(".")
    imported = 0
    if ext == "pdf":
        reader = PdfReader(str(path))
        parts = []
        for index, page in enumerate(reader.pages, start=1):
            page_text = page.extract_text() or ""
            if page_text.strip():
                parts.append(f"--- Página {index} ---\n{page_text.strip()}")
        return "\n\n".join(parts) or "Não foi possível extrair texto deste PDF.", imported
    if ext in {"xlsx", "xlsm"}:
        workbook = load_workbook(path, data_only=True, read_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"Planilha: {sheet.title}")
            rows = list(sheet.iter_rows(values_only=True))
            if import_controls and rows:
                imported += import_controls_from_sheet(rows)
            for row in rows[:80]:
                values = [str(value) if value is not None else "" for value in row]
                lines.append(" | ".join(values).strip())
        return "\n".join(lines), imported
    if ext == "csv":
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        if import_controls and rows:
            imported = import_controls_from_sheet(rows)
        return text[:50000], imported
    if ext == "txt":
        return path.read_text(encoding="utf-8", errors="replace")[:50000], imported
    return "Extração automática disponível para PDF, Excel, CSV e TXT.", imported


def import_controls_from_sheet(rows: list[tuple[Any, ...] | list[Any]]) -> int:
    if not rows:
        return 0
    headers = [str(value or "").strip().lower() for value in rows[0]]
    aliases = {
        "titulo": ["titulo", "título", "controle", "nome", "nome do controle"],
        "descricao": ["descricao", "descrição"],
        "categoria": ["categoria"],
        "tipo": ["tipo"],
        "periodicidade": ["periodicidade"],
        "criticidade": ["criticidade", "grau de criticidade"],
        "status": ["status"],
        "prazo": ["prazo", "data prevista", "data de revisão", "data revisao"],
    }

    def find_column(names: list[str]) -> int | None:
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    columns = {key: find_column(names) for key, names in aliases.items()}
    if columns["titulo"] is None:
        return 0
    people = get_choice_options("pessoas")
    default_owner = int(people[0][0]) if people else None
    if not default_owner:
        return 0
    imported = 0
    for row in rows[1:]:
        values = list(row)
        title = values[columns["titulo"]] if columns["titulo"] is not None and columns["titulo"] < len(values) else None
        if not title:
            continue
        payload = {
            "titulo": str(title).strip(),
            "descricao": str(values[columns["descricao"]]).strip() if columns["descricao"] is not None and columns["descricao"] < len(values) and values[columns["descricao"]] else "",
            "categoria": str(values[columns["categoria"]]).strip() if columns["categoria"] is not None and columns["categoria"] < len(values) and values[columns["categoria"]] else "Importado",
            "tipo": str(values[columns["tipo"]]).strip() if columns["tipo"] is not None and columns["tipo"] < len(values) and values[columns["tipo"]] else "Segurança",
            "eixo": None,
            "area_responsavel": "Importado",
            "responsavel_id": default_owner,
            "periodicidade": str(values[columns["periodicidade"]]).strip() if columns["periodicidade"] is not None and columns["periodicidade"] < len(values) and values[columns["periodicidade"]] else "Anual",
            "status": str(values[columns["status"]]).strip() if columns["status"] is not None and columns["status"] < len(values) and values[columns["status"]] else "Não iniciado",
            "criticidade": str(values[columns["criticidade"]]).strip() if columns["criticidade"] is not None and columns["criticidade"] < len(values) and values[columns["criticidade"]] else "Média",
            "data_criacao": current_date().isoformat(),
            "data_atualizacao": current_date().isoformat(),
            "data_revisao": None,
            "prazo": None,
            "observacoes": "Importado automaticamente de planilha.",
        }
        payload["eixo"] = area_title_for_control(payload["titulo"], payload["categoria"])
        get_db().execute(
            """
            INSERT INTO controles
                (titulo, descricao, categoria, tipo, eixo, area_responsavel, responsavel_id,
                 periodicidade, status, criticidade, data_criacao, data_atualizacao,
                 data_revisao, prazo, observacoes)
            VALUES
                (:titulo, :descricao, :categoria, :tipo, :eixo, :area_responsavel, :responsavel_id,
                 :periodicidade, :status, :criticidade, :data_criacao, :data_atualizacao,
                 :data_revisao, :prazo, :observacoes)
            """,
            payload,
        )
        imported += 1
    get_db().commit()
    return imported


REPORT_TYPES = {
    "geral": "Relatório Geral de Segurança da Informação",
    "controles": "Relatório de Controles",
    "riscos": "Relatório de Riscos",
    "incidentes": "Relatório de Incidentes",
    "melhorias": "Relatório de Melhorias",
    "evidencias": "Relatório de Evidências",
    "executivo": "Relatório Executivo",
}


@app.route("/relatorios")
@permission_required("relatorios", "view")
def reports():
    kind = request.args.get("tipo", "executivo")
    report = build_report(kind)
    control_area = "melhorias" if kind == "melhorias" else None
    if kind == "evidencias":
        control_area = "arquivos"
    return render_template(
        "reports.html",
        report=report,
        report_types=REPORT_TYPES,
        filters=request.args,
        people=get_choice_options("pessoas"),
        control_area=control_area,
    )


@app.route("/relatorios/exportar/<fmt>")
@permission_required("relatorios", "export")
def report_export(fmt: str):
    kind = request.args.get("tipo", "executivo")
    report = build_report(kind)
    if fmt == "csv":
        response = export_report_csv(report)
    elif fmt == "xlsx":
        response = export_report_xlsx(report)
    elif fmt == "pdf":
        response = export_report_pdf(report)
    else:
        abort(404)
    audit("exportar_relatorio", "relatorios", None, f"{report['title']} em {fmt.upper()}")
    return response


def report_filter_sql(alias: str, field_map: dict[str, str]) -> tuple[str, list[Any]]:
    clauses = []
    params: list[Any] = []
    for request_key, column in field_map.items():
        if request_key in {"inicio", "fim"}:
            continue
        value = request.args.get(request_key, "").strip()
        if value:
            clauses.append(f"{column} = ?")
            params.append(value)
    if request.args.get("inicio") and "inicio" in field_map:
        clauses.append(f"{field_map['inicio']} >= ?")
        params.append(request.args["inicio"])
    if request.args.get("fim") and "fim" in field_map:
        clauses.append(f"{field_map['fim']} <= ?")
        params.append(request.args["fim"])
    return (" AND " + " AND ".join(clauses)) if clauses else "", params


def build_report(kind: str) -> dict[str, Any]:
    if kind not in REPORT_TYPES:
        kind = "executivo"
    db = get_db()
    title = REPORT_TYPES[kind]
    if kind in {"geral", "executivo"}:
        data = dashboard_data()
        rows = [{"indicador": card["label"], "valor": card["value"], "observacao": card.get("hint", "")} for card in data["cards"]]
        if kind == "executivo":
            rows.extend({"indicador": f"Ponto de atenção: {alert['title']}", "valor": "", "observacao": alert["text"]} for alert in data["alerts"][:6])
            rows.extend({"indicador": f"Recomendação: {item['title']}", "valor": "", "observacao": item["text"]} for item in data["improvements"][:6])
        return {"kind": kind, "title": title, "columns": [("indicador", "Indicador"), ("valor", "Valor"), ("observacao", "Observação")], "rows": rows}

    if kind == "controles":
        where, params = report_filter_sql(
            "c",
            {"status": "c.status", "categoria": "c.categoria", "responsavel_id": "c.responsavel_id", "setor": "c.area_responsavel", "inicio": "c.data_criacao", "fim": "c.data_criacao"},
        )
        rows = db.execute(
            f"""
            SELECT c.titulo, c.categoria, c.tipo, p.nome AS responsavel, c.status,
                   c.periodicidade, c.data_revisao, c.prazo,
                   (SELECT COUNT(*) FROM arquivos a WHERE a.controle_id = c.id AND a.ativo = 1) AS evidencias
            FROM controles c
            LEFT JOIN pessoas p ON p.id = c.responsavel_id
            WHERE c.ativo = 1 {where}
            ORDER BY c.categoria, c.titulo
            """,
            params,
        ).fetchall()
        return {"kind": kind, "title": title, "columns": [("titulo", "Controle"), ("categoria", "Categoria"), ("tipo", "Tipo"), ("responsavel", "Responsável"), ("status", "Status"), ("periodicidade", "Periodicidade"), ("data_revisao", "Revisão"), ("prazo", "Prazo"), ("evidencias", "Evidências")], "rows": rows}

    if kind == "riscos":
        where, params = report_filter_sql("r", {"status": "r.status", "categoria": "r.categoria", "responsavel_id": "r.responsavel_id"})
        rows = db.execute(
            f"""
            SELECT r.titulo, r.categoria, c.titulo AS controle, r.probabilidade, r.impacto,
                   r.pontuacao, r.nivel_risco, p.nome AS responsavel, r.status, r.plano_tratamento, r.prazo
            FROM riscos r
            LEFT JOIN controles c ON c.id = r.controle_id
            LEFT JOIN pessoas p ON p.id = r.responsavel_id
            WHERE r.ativo = 1 {where}
            ORDER BY r.pontuacao DESC, r.prazo
            """,
            params,
        ).fetchall()
        return {"kind": kind, "title": title, "columns": [("titulo", "Risco"), ("categoria", "Categoria"), ("controle", "Controle"), ("probabilidade", "Prob."), ("impacto", "Impacto"), ("pontuacao", "Pontuação"), ("nivel_risco", "Nível"), ("responsavel", "Responsável"), ("status", "Status"), ("plano_tratamento", "Plano"), ("prazo", "Prazo")], "rows": rows}

    if kind == "incidentes":
        where, params = report_filter_sql("i", {"status": "i.status", "responsavel_id": "i.responsavel_id", "setor": "i.area_afetada", "inicio": "i.data_abertura", "fim": "i.data_abertura"})
        rows = db.execute(
            f"""
            SELECT i.numero, i.titulo, i.tipo, i.gravidade, i.ativo_afetado, i.area_afetada,
                   p.nome AS responsavel, i.status, i.data_abertura, i.data_encerramento, i.causa, i.acoes_tomadas
            FROM incidentes i
            LEFT JOIN pessoas p ON p.id = i.responsavel_id
            WHERE i.ativo = 1 {where}
            ORDER BY i.data_abertura DESC
            """,
            params,
        ).fetchall()
        return {"kind": kind, "title": title, "columns": [("numero", "Número"), ("titulo", "Incidente"), ("tipo", "Tipo"), ("gravidade", "Gravidade"), ("ativo_afetado", "Ativo"), ("area_afetada", "Área"), ("responsavel", "Responsável"), ("status", "Status"), ("data_abertura", "Abertura"), ("data_encerramento", "Encerramento"), ("causa", "Causa"), ("acoes_tomadas", "Ações tomadas")], "rows": rows}

    if kind == "melhorias":
        alerts, improvements = generate_alerts_and_improvements(db)
        rows = [{"tipo": "Ponto de atenção", "titulo": item["title"], "descricao": item["text"], "prioridade": item.get("level", "")} for item in alerts]
        rows.extend({"tipo": "Melhoria sugerida", "titulo": item["title"], "descricao": item["text"], "prioridade": ""} for item in improvements)
        return {"kind": kind, "title": title, "columns": [("tipo", "Tipo"), ("titulo", "Título"), ("descricao", "Descrição"), ("prioridade", "Prioridade")], "rows": rows}

    rows = db.execute(
        """
        SELECT a.nome_original, a.tipo, a.classificacao, u.nome AS responsavel_envio, a.criado_em,
               a.usuario_upload_id, c.titulo AS controle, r.titulo AS risco, i.numero AS incidente, ac.titulo AS acao
        FROM arquivos a
        LEFT JOIN usuarios u ON u.id = a.usuario_upload_id
        LEFT JOIN controles c ON c.id = a.controle_id
        LEFT JOIN riscos r ON r.id = a.risco_id
        LEFT JOIN incidentes i ON i.id = a.incidente_id
        LEFT JOIN acoes ac ON ac.id = a.acao_id
        WHERE a.ativo = 1
        ORDER BY a.criado_em DESC
        """
    ).fetchall()
    rows = [row for row in rows if can_access_file(row, "view")] if g.get("user") else rows
    return {"kind": kind, "title": title, "columns": [("nome_original", "Documento"), ("tipo", "Tipo"), ("classificacao", "Classificação"), ("responsavel_envio", "Enviado por"), ("criado_em", "Data"), ("controle", "Controle"), ("risco", "Risco"), ("incidente", "Incidente"), ("acao", "Ação")], "rows": rows}


def cell_value(row: Any, key: str) -> str:
    value = row.get(key) if isinstance(row, dict) else row[key]
    if key.startswith("data") or key in {"prazo", "criado_em", "atualizado_em"}:
        return format_date(value)
    if key in {"probabilidade", "impacto"}:
        return {1: "Baixa", 2: "Média", 3: "Alta", "1": "Baixa", "2": "Média", "3": "Alta"}.get(value, str(value or ""))
    return str(value or "")


def report_filename(report: dict[str, Any], ext: str) -> str:
    slug = report["kind"].replace("_", "-")
    return f"pgsi-{slug}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.{ext}"


def export_report_csv(report: dict[str, Any]) -> Response:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([label for _, label in report["columns"]])
    for row in report["rows"]:
        writer.writerow([cell_value(row, key) for key, _ in report["columns"]])
    data = output.getvalue().encode("utf-8-sig")
    return Response(
        data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={report_filename(report, 'csv')}"},
    )


def export_report_xlsx(report: dict[str, Any]) -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"
    sheet.append([label for _, label in report["columns"]])
    for row in report["rows"]:
        sheet.append([cell_value(row, key) for key, _ in report["columns"]])
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
        sheet.column_dimensions[letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=report_filename(report, "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def export_report_pdf(report: dict[str, Any]) -> Response:
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [Paragraph(report["title"], styles["Title"]), Spacer(1, 12)]
    table_data = [[Paragraph(label, styles["BodyText"]) for _, label in report["columns"]]]
    for row in report["rows"][:500]:
        table_data.append([Paragraph(cell_value(row, key)[:450], styles["BodyText"]) for key, _ in report["columns"]])
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7F0EE")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#18211F")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C5C0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9F8")]),
            ]
        )
    )
    elements.append(table)
    document.build(elements)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=report_filename(report, "pdf"), mimetype="application/pdf")


@app.route("/logs")
@permission_required("logs", "view")
def logs():
    rows = get_db().execute(
        """
        SELECT l.*, u.nome AS usuario_nome
        FROM logs_auditoria l
        LEFT JOIN usuarios u ON u.id = l.usuario_id
        ORDER BY l.id DESC
        LIMIT 250
        """
    ).fetchall()
    return render_template("logs.html", rows=rows)


@app.route("/backup", methods=["GET", "POST"])
@permission_required("backup", "view")
def backup():
    if request.method == "POST":
        if not has_permission("backup", "create"):
            abort(403)
        filename = f"pgsi-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.sqlite3"
        destination = BACKUP_DIR / filename
        shutil.copy2(DATABASE, destination)
        audit("backup", "database", None, filename)
        flash("Backup criado com sucesso.", "success")
        return redirect(url_for("backup"))
    files = [
        {
            "name": path.name,
            "size": path.stat().st_size,
            "modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        }
        for path in sorted(BACKUP_DIR.glob("*.sqlite3"), key=lambda item: item.stat().st_mtime, reverse=True)
    ]
    return render_template("backup.html", files=files)


@app.route("/backup/<name>")
@permission_required("backup", "view")
def backup_download(name: str):
    safe_name = secure_filename(name)
    path = BACKUP_DIR / safe_name
    if not path.exists():
        abort(404)
    return send_file(path, as_attachment=True, download_name=safe_name)


@app.errorhandler(403)
def forbidden(_: Exception):
    return render_template("error.html", code=403, message="Você não tem permissão para acessar esta área."), 403


@app.errorhandler(404)
def not_found(_: Exception):
    return render_template("error.html", code=404, message="Registro ou página não encontrada."), 404


def _verify_api_token() -> bool:
    token = os.environ.get("PGSI_API_TOKEN", "")
    if not token:
        return False
    auth = request.headers.get("Authorization", "")
    return auth == f"Bearer {token}"


@app.route("/api/incidentes", methods=["POST"])
def api_create_incidente():
    if not _verify_api_token():
        return {"erro": "Não autorizado"}, 401

    data = request.get_json(silent=True)
    if not data:
        return {"erro": "JSON inválido"}, 400

    titulo = (data.get("titulo") or "").strip()
    if not titulo:
        return {"erro": "Campo 'titulo' obrigatório"}, 400

    tipo = data.get("tipo", "Malware")
    gravidade = data.get("gravidade", "Alta")
    ativo_afetado = data.get("ativo_afetado", "")
    area_afetada = data.get("area_afetada", "")
    descricao = data.get("descricao", "")
    causa = data.get("causa", "")
    numero = data.get("numero") or f"INC-{datetime.now().strftime('%Y-%m%d-%H%M%S')}"
    data_abertura = data.get("data_abertura") or date.today().isoformat()

    tipos_validos = ["Indisponibilidade de sistema", "Vazamento de dados", "Acesso indevido",
                     "Malware", "Phishing", "Falha de backup", "Falha de rede",
                     "Perda de equipamento", "Outro"]
    gravidades_validas = ["Baixa", "Média", "Alta", "Crítica"]

    if tipo not in tipos_validos:
        tipo = "Malware"
    if gravidade not in gravidades_validas:
        gravidade = "Alta"

    db = get_db()

    numero_existe = db.execute("SELECT id FROM incidentes WHERE numero = ?", (numero,)).fetchone()
    if numero_existe:
        return {"erro": f"Incidente {numero} já existe", "id": numero_existe["id"]}, 409

    cursor = db.execute(
        """
        INSERT INTO incidentes
            (numero, titulo, descricao, tipo, gravidade, ativo_afetado,
             area_afetada, status, data_abertura, causa, criado_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'Aberto', ?, ?, CURRENT_TIMESTAMP)
        """,
        (numero, titulo, descricao, tipo, gravidade, ativo_afetado,
         area_afetada, data_abertura, causa),
    )
    db.commit()
    incidente_id = cursor.lastrowid

    db.execute(
        """
        INSERT INTO logs_auditoria (usuario_id, acao, tabela_afetada, registro_id, ip, detalhes)
        VALUES (NULL, 'criar_api', 'incidentes', ?, ?, ?)
        """,
        (incidente_id, request.remote_addr, f"Incidente criado via API: {titulo}"),
    )
    db.commit()

    return {"sucesso": True, "id": incidente_id, "numero": numero}, 201


with app.app_context():
    init_db()


if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1")
